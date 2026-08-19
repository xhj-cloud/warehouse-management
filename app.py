"""
仓库管理系统 - Flask 主应用
"""

import os
import sys
import json
import time
import uuid
import decimal
import traceback
import logging
import pymysql
import pandas as pd
from datetime import datetime, date, timedelta
import base64
import hmac
from flask import (
    Flask, request, jsonify, render_template, send_from_directory, send_file,
    make_response, session, redirect, url_for, has_request_context,
)
from functools import wraps
from flask.json.provider import DefaultJSONProvider

# 错误日志文件
ERROR_LOG = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'error.log')
logging.basicConfig(
    filename=ERROR_LOG,
    level=logging.ERROR,
    format='%(asctime)s [%(levelname)s] %(message)s',
)
_logger = logging.getLogger(__name__)
from werkzeug.utils import secure_filename

from config import (
    UPLOAD_FOLDER, ALLOWED_EXTENSIONS, SECRET_KEY, MAX_CONTENT_LENGTH,
    AUTH_USER, AUTH_PASSWORD, AUTH_DISABLED,
)
from models import (
    db, CategoryModel, ProductModel, InventoryModel,
    TransactionModel, ExcelUploadModel, StatsModel,
    SupplierModel, CustomerModel, AuditLog,
)
from models import UserModel, seed_admin_if_empty
from ai_service import ai_service


class CustomJSONProvider(DefaultJSONProvider):
    """自定义 JSON 提供器，处理 Decimal / date 等类型"""
    def default(self, obj):
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        if isinstance(obj, bytes):
            return obj.decode('utf-8', errors='replace')
        return super().default(obj)


app = Flask(__name__)
app.json = CustomJSONProvider(app)
app.secret_key = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH
# 登录态保持 7 天（持久 cookie）：刷新页面/重开浏览器不再需要重新登录
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# 确保上传目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ==========================================
#  认证：Session 登录（数据库用户表 + 环境变量超级管理员兜底）
# ==========================================
# users 表为空时引导默认 admin/admin123（首次部署自动创建，不依赖硬编码哈希进 SQL）
seed_admin_if_empty()

# 认证始终启用（除非显式 DISABLE_AUTH）。
# - 登录页 /login 提交用户名密码 → 校验 users 表 → 写入 session（7 天有效）
# - 环境变量 AUTH_USER/AUTH_PASSWORD 作为超级管理员兜底（不受用户表影响）
_AUTH_ENABLED = not AUTH_DISABLED

if _AUTH_ENABLED:
    print("  认证已启用：访问需登录（可用账户管理页创建账户）")
else:
    print("  警告：已通过 DISABLE_AUTH 显式关闭认证（仅限本地联调，请勿用于线上/内网）")


def _validate_credentials(username, password):
    """校验用户名+密码：users 表优先，环境变量超级管理员兜底。成功返回 (ok, role)。"""
    # 1) 数据库用户表校验
    user = UserModel.authenticate(username, password)
    if user:
        return True, user.get('role') or 'user'
    # 2) 环境变量超级管理员兜底（常量时间比较）
    if AUTH_PASSWORD and hmac.compare_digest(str(username), str(AUTH_USER)) \
            and hmac.compare_digest(password or '', str(AUTH_PASSWORD)):
        return True, 'admin'
    return False, None


def _current_username():
    """返回当前已登录用户名；未登录返回 None。"""
    return session.get('username')


def _op(default='系统'):
    """操作者：优先当前登录账号，未登录/无请求上下文时回退 default。

    出入库流水/审计日志的 operator 一律以此为准，不再信任前端传入的 operator，
    确保「记录操作者 = 登录账号」（防止伪造他人操作）。
    """
    if has_request_context():
        u = session.get('username')
        if u:
            return u
    return default


def _is_admin():
    """当前登录用户是否管理员（登录时写入 session 的 role=admin）。"""
    return session.get('role') == 'admin'


def _require_login():
    """判断当前请求是否已登录（未启用认证时视为已登录）。"""
    if not _AUTH_ENABLED:
        return True
    return bool(session.get('username'))


@app.route('/login')
def login_page():
    """登录页（GET）。"""
    if _require_login():
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    """登录接口：校验用户名密码，写入 session（7 天有效）。"""
    try:
        data = request.json or {}
        username = _clean_cell(data.get('username'))
        password = data.get('password') or ''
        if not username or not password:
            return jsonify({'success': False, 'error': '请输入用户名和密码'}), 400
        ok, role = _validate_credentials(username, password)
        if not ok:
            return jsonify({'success': False, 'error': '用户名或密码错误'}), 401
        session.permanent = True          # 7 天持久 cookie，刷新/重开页面保持登录
        session['username'] = username
        session['role'] = role
        return jsonify({'success': True, 'data': {'username': username, 'role': role}})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """退出登录：清空 session。"""
    session.clear()
    return jsonify({'success': True})


@app.before_request
def _authenticate():
    """所有页面与 /api/*、/uploads/* 均需登录（登录页与静态资源除外）。"""
    path = request.path
    # 登录页 / 登录接口 / 静态资源 / favicon 公开（登录接口必须在白名单，否则无法登录）
    if path in ('/login', '/api/login') or path.startswith('/static/') or path == '/favicon.ico':
        return None
    if _require_login():
        return None
    # 未登录：API 返回 401 JSON，页面重定向到登录页
    if path.startswith('/api/') or path.startswith('/uploads/'):
        return jsonify({'success': False, 'error': '未登录', 'code': 'AUTH_REQUIRED'}), 401
    return redirect(url_for('login_page'))


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _clean_cell(value):
    """把 pandas 空单元格（NaN/None/空串）统一归一为 ''，其余转字符串去首尾空白。"""
    if value is None:
        return ''
    if isinstance(value, float):
        import math
        if math.isnan(value):
            return ''
    s = str(value).strip()
    return '' if s.lower() == 'nan' else s


def _to_int(value):
    """宽容的整数转换：兼容 '10'、10、10.0、'10.5'（向下取整）；NaN/None/无法解析返回 None。"""
    if value is None or isinstance(value, bool):
        return None
    try:
        f = float(str(value).strip())
    except (ValueError, TypeError):
        return None
    if f != f or f in (float('inf'), float('-inf')):  # NaN / ±Inf
        return None
    return int(f)


def _to_float(value, default=0.0):
    """宽容的浮点转换；NaN/None/无法解析返回 default。"""
    if value is None or isinstance(value, bool):
        return default
    try:
        f = float(str(value).strip())
    except (ValueError, TypeError):
        return default
    return f if f == f and f not in (float('inf'), float('-inf')) else default


def _read_upload_df(filepath, ext):
    """按扩展名读取上传的表格文件（含 CSV 编码回退与旧版 .xls 支持）。"""
    if ext == 'csv':
        try:
            return pd.read_csv(filepath, encoding='utf-8-sig')
        except UnicodeDecodeError:
            return pd.read_csv(filepath, encoding='gbk')
    if ext == 'xls':
        return pd.read_excel(filepath, engine='xlrd')
    return pd.read_excel(filepath, engine='openpyxl')


def _apply_inventory_change(prod_id, new_qty, location='', min_stock=None, max_stock=None, operator=None):
    """以「出入库流水」的方式把库存调整到目标数量（保留审计），并同步库位/阈值。

    必须在 ``with db.transaction():`` 内调用，保证数量变更与流水写入原子化。
    用 SELECT ... FOR UPDATE 锁定目标行，杜绝并发「设为绝对值」时的读-改-写竞态
    （否则两个请求都按旧值算 delta，最后落库值 ≠ 目标值）。
    """
    operator = operator or _op()   # 操作者 = 登录账号
    # FOR UPDATE 行锁：在事务内锁住该行，后到的并发请求会阻塞到前一个提交，
    # 读到的必然是最新值，从而消除"先读再算 delta"的丢失更新。
    old_inv = db.query_one("SELECT quantity FROM inventory WHERE product_id=%s FOR UPDATE", (prod_id,))
    old_qty = old_inv['quantity'] if old_inv else 0
    delta = new_qty - old_qty
    if delta > 0:
        InventoryModel.stock_in(prod_id, delta, operator=operator, notes='手动调整库存')
    elif delta < 0:
        InventoryModel.stock_out(prod_id, -delta, operator=operator, notes='手动调整库存')
    db.execute(
        "UPDATE inventory SET location=%s, min_stock=%s, max_stock=%s WHERE product_id=%s",
        (location or '',
         min_stock if min_stock is not None else 0,
         max_stock if max_stock is not None else 9999,
         prod_id))


def _safe_filename_part(name, default='客户'):
    """只保留中文/字母/数字/下划线/连字符，其余替换为下划线；用于出库单文件名，防路径穿越。"""
    import re
    s = re.sub(r'[^\w\u4e00-\u9fff-]+', '_', str(name or '').strip())
    return s.strip('_') or default


def _unique_sku(base):
    """基于 base 生成不冲突的 SKU：已存在则追加 -2/-3...

    AI 识别路径用 name[:6] 兜底生成 SKU，中文商品名前 6 字常相同，
    直接复用会命中别的商品、把库存并入错误记录。
    """
    candidate = base
    n = 1
    while ProductModel.get_by_sku(candidate):
        n += 1
        candidate = f'{base}-{n}'
    return candidate


def _ai_batch_no(prefix):
    """生成 AI 操作唯一批次号（前缀-时间戳-随机后缀）。

    旧实现用常量批次号（'AI视觉识别'/'AI智能导入'/'AI操作'），多次导入在
    批次视图里合并成一坨、无法追溯；现在每次调用独立成批，与 Excel-{upload_id} 对齐。
    """
    return f'{prefix}-{datetime.now().strftime("%Y%m%d%H%M%S")}-{uuid.uuid4().hex[:6]}'


def _cleanup_old_order_files(keep_days=7):
    """清理 uploads/ 中过期的出库单 Excel，防止磁盘无限增长（只动 出库单_*.xlsx）。"""
    try:
        cutoff = time.time() - keep_days * 86400
        for fn in os.listdir(UPLOAD_FOLDER):
            if not (fn.startswith('出库单_') and fn.endswith('.xlsx')):
                continue
            fp = os.path.join(UPLOAD_FOLDER, fn)
            try:
                if os.path.isfile(fp) and os.path.getmtime(fp) < cutoff:
                    os.remove(fp)
            except OSError:
                pass
    except OSError:
        pass


# ==========================================
#  账户管理 API（仅管理员可操作）
# ==========================================
@app.route('/api/auth/me')
def api_auth_me():
    """返回当前登录用户信息（供前端显示当前账户与是否为管理员）。"""
    username = _current_username()
    if not _require_login():
        return jsonify({'success': False, 'error': '未登录', 'code': 'AUTH_REQUIRED'}), 401
    return jsonify({'success': True, 'data': {
        'username': username,
        'role': session.get('role'),
        'is_admin': _is_admin(),
    }})


@app.route('/api/users')
def api_users_list():
    """列出所有账户（仅管理员）。"""
    if not _is_admin():
        return jsonify({'success': False, 'error': '无权限，仅管理员可查看账户'}), 403
    try:
        users = UserModel.get_all()
        return jsonify({'success': True, 'data': users})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/users', methods=['POST'])
def api_users_create():
    """创建账户（仅管理员）。body: {username, password, role}"""
    if not _is_admin():
        return jsonify({'success': False, 'error': '无权限，仅管理员可创建账户'}), 403
    try:
        data = request.json or {}
        username = _clean_cell(data.get('username'))
        password = data.get('password') or ''
        role = _clean_cell(data.get('role')) or 'user'
        if not username:
            return jsonify({'success': False, 'error': '用户名不能为空'}), 400
        if len(password) < 6:
            return jsonify({'success': False, 'error': '密码至少 6 位'}), 400
        if role not in ('admin', 'user'):
            role = 'user'
        # 用户名规范化：小写去空格
        username = username.strip().lower()
        if UserModel.get_by_username(username):
            return jsonify({'success': False, 'error': f'用户名 {username} 已存在'}), 400
        uid = UserModel.create(username, password, role)
        AuditLog.log('create', 'users', uid,
                     new_data={'username': username, 'role': role},
                     operator=_current_username())
        return jsonify({'success': True, 'id': uid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/users/<int:uid>', methods=['DELETE'])
def api_users_delete(uid):
    """删除账户（仅管理员；不能删除自己或最后一个管理员）。"""
    if not _is_admin():
        return jsonify({'success': False, 'error': '无权限，仅管理员可删除账户'}), 403
    try:
        target = db.query_one("SELECT * FROM users WHERE id=%s", (uid,))
        if not target:
            return jsonify({'success': False, 'error': '账户不存在'}), 404
        # 不能删除自己
        if target['username'] == _current_username():
            return jsonify({'success': False, 'error': '不能删除当前登录的账户'}), 400
        # 不能删除最后一个管理员
        if target.get('role') == 'admin' and UserModel.count_admins() <= 1:
            return jsonify({'success': False, 'error': '不能删除最后一个管理员账户'}), 400
        UserModel.delete(uid)
        AuditLog.log('delete', 'users', uid,
                     new_data={'username': target['username']},
                     operator=_current_username())
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  审计日志 API
# ==========================================
@app.route('/api/audit-log')
def api_audit_log():
    try:
        limit = int(request.args.get('limit', 200))
        table = request.args.get('table', '')
        if table:
            logs = AuditLog.get_by_table(table, limit)
        else:
            logs = AuditLog.get_recent(limit)
        return jsonify({'success': True, 'data': logs})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  库存批次 API
# ==========================================
@app.route('/api/inventory/batches')
def api_inventory_batches():
    """获取每个商品按供应商/进价的进货批次汇总"""
    try:
        batches = db.query("""
            SELECT p.id AS product_id, p.name AS product_name, p.sku,
                   p.unit, p.unit_price AS avg_price, i.quantity AS total_qty,
                   t.batch_no, t.unit_price AS batch_price,
                   s.name AS supplier_name,
                   SUM(CASE WHEN t.type='in' THEN t.quantity ELSE 0 END) -
                   SUM(CASE WHEN t.type='out' THEN t.quantity ELSE 0 END) AS net_qty
            FROM products p
            JOIN inventory i ON p.id = i.product_id
            LEFT JOIN transactions t ON t.product_id = p.id
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            GROUP BY p.id, p.name, p.sku, p.unit, p.unit_price, i.quantity,
                     s.name, t.batch_no, t.unit_price
            HAVING net_qty > 0
            ORDER BY p.name, MAX(t.id) DESC
        """)
        return jsonify({'success': True, 'data': batches})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  出库单 Excel 导出
# ==========================================
@app.route('/api/order/export-inventory', methods=['POST'])
def api_export_inventory():
    """导出库存清单 Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    try:
        data = request.get_json(silent=True) or {}
        items = data.get('items', [])
        if not isinstance(items, list):
            return jsonify({'success': False, 'error': '参数错误：items 必须是数组'}), 400
        if any(not isinstance(it, dict) for it in items):
            return jsonify({'success': False, 'error': '参数错误：items 中存在非对象条目'}), 400

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '库存清单'
        hfill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        hfont = Font(color='FFFFFF', bold=True, size=11)
        bd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ca = Alignment(horizontal='center', vertical='center')
        ws['A1'] = '库存清单'; ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = ca
        ws.merge_cells('A1:K1')
        ws['A2'] = '导出日期: ' + datetime.now().strftime('%Y-%m-%d')
        headers = ['商品名称','SKU','分类','供应商','单位','规格','库存','最低库存','最高库存','库位','最近进价']
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=ci, value=h); c.fill = hfill; c.font = hfont; c.alignment = ca; c.border = bd
        for idx, item in enumerate(items, 1):
            r = 4 + idx
            for ci, k in enumerate(['product_name','sku','category_name','supplier_name','unit','','quantity','min_stock','max_stock','location','latest_price'], 1):
                v = item.get(k, '') if k else (item.get('specification', '') or '')
                if k == 'latest_price' and v:
                    try:
                        if float(v) > 0: v = '¥' + str(v)
                    except (TypeError, ValueError):
                        pass  # 非法价格原样输出，不让整单导出失败
                c = ws.cell(row=r, column=ci, value=v if v is not None else '')
                c.border = bd; c.alignment = ca
        for i, w in enumerate([18,15,12,12,8,15,8,10,10,10,12], 1):
            ws.column_dimensions[chr(64+i)].width = w
        out = BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='库存清单_' + datetime.now().strftime('%Y%m%d_%H%M%S') + '.xlsx')
    except Exception as e:
        _logger.error(f"库存清单导出错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'导出失败: {str(e)}'}), 500


@app.route('/api/order/export', methods=['POST'])
def api_order_export():
    """生成出库单 Excel 文件"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO

    try:
        data = request.get_json(silent=True)
        if not isinstance(data, dict):
            return jsonify({'success': False, 'error': '请求体必须是 JSON 对象'}), 400
        items = data.get('items', [])
        if not isinstance(items, list):
            return jsonify({'success': False, 'error': '参数错误：items 必须是数组'}), 400
        if any(not isinstance(it, dict) for it in items):
            return jsonify({'success': False, 'error': '参数错误：items 中存在非对象条目'}), 400

        customer = data.get('customer', '') or ''
        operator = data.get('operator', '') or ''
        keeper = data.get('keeper', '') or ''
        warehouse_no = data.get('warehouse', '') or ''
        now_str = datetime.now().strftime('%Y-%m-%d')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '出库单'

        # 样式
        title_font = Font(size=16, bold=True)
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        label_font = Font(size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))
        center = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # 标题
        ws.merge_cells('A1:G1')
        ws['A1'] = '出 库 单'
        ws['A1'].font = title_font
        ws['A1'].alignment = center

        # 日期和客户（标题下方一行，无空行）
        ws['A2'] = '日期:'; ws['A2'].font = label_font; ws['A2'].alignment = right_align
        ws['B2'] = now_str; ws['B2'].font = label_font; ws['B2'].alignment = left_align
        ws['D2'] = '客户:'; ws['D2'].font = label_font; ws['D2'].alignment = right_align
        ws['E2'] = customer; ws['E2'].font = label_font; ws['E2'].alignment = left_align

        # 表头（第3行开始，无空行）
        headers = ['序号', '商品名称', '规格', 'SKU', '数量', '单价(元)', '金额(元)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill; cell.font = header_font
            cell.alignment = center; cell.border = thin_border

        # 数据行（宽容转换：非法价格/数量按 0 计，不让整单导出失败）
        total = 0
        for idx, item in enumerate(items, 1):
            row = 3 + idx
            price = _to_float(item.get('sale_price'))
            qty = _to_int(item.get('quantity')) or 0
            sub = round(price * qty, 2)
            total += sub
            values = [idx, item.get('name', ''), item.get('specification', ''),
                      item.get('sku', ''), qty, price, sub]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center if col != 6 and col != 7 else right_align

        # 合计行（紧接数据）
        sum_row = 3 + len(items) + 1
        ws.merge_cells(f'A{sum_row}:E{sum_row}')
        ws.cell(row=sum_row, column=1, value='合计').font = Font(bold=True, size=12)
        ws.cell(row=sum_row, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=sum_row, column=6, value='').border = thin_border
        total_cell = ws.cell(row=sum_row, column=7, value=total)
        total_cell.font = Font(bold=True, size=12, color='FF0000')
        total_cell.border = thin_border; total_cell.alignment = right_align

        # 经办人、库管、仓库编号（合计下方，无空行）
        bot = sum_row + 1
        ws.cell(row=bot, column=1, value='经办人:').font = label_font
        ws.cell(row=bot, column=1).alignment = right_align
        ws.cell(row=bot, column=2, value=operator).font = label_font
        ws.cell(row=bot, column=2).alignment = left_align
        ws.cell(row=bot, column=4, value='库管:').font = label_font
        ws.cell(row=bot, column=4).alignment = right_align
        ws.cell(row=bot, column=5, value=keeper).font = label_font
        ws.cell(row=bot, column=5).alignment = left_align
        ws.cell(row=bot+1, column=1, value='仓库编号:').font = label_font
        ws.cell(row=bot+1, column=1).alignment = right_align
        ws.cell(row=bot+1, column=2, value=warehouse_no).font = label_font
        ws.cell(row=bot+1, column=2).alignment = left_align

        # 列宽
        widths = [6, 22, 20, 15, 8, 12, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 输出（文件名净化 + 精确到秒的时间戳，防路径穿越与同日覆盖）
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"出库单_{_safe_filename_part(customer)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        _logger.error(f"出库单导出错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'导出失败: {str(e)}'}), 500


# ==========================================
#  入库单视觉识别
# ==========================================
@app.route('/api/ai/inbound-recognize', methods=['POST'])
def api_inbound_recognize():
    """AI 识别入库单图片"""
    try:
        data = request.json
        image_b64 = data.get('image', '')
        if not image_b64:
            return jsonify({'success': False, 'error': '请上传图片'}), 400

        items = ai_service.recognize_inbound_image(image_b64)
        if not isinstance(items, list):
            return jsonify({'success': False, 'error': 'AI 返回格式异常，请重试或使用更清晰的图片'}), 400

        imported = []
        skipped = []
        # 本次识别独立成批（旧常量 'AI视觉识别' 会让多次导入在批次视图里合并）
        batch_no = _ai_batch_no('AI识别')
        with db.transaction():
            for item in items:
                if not isinstance(item, dict):
                    skipped.append('忽略非对象条目')
                    continue
                name = _clean_cell(item.get('name'))
                qty = _to_int(item.get('quantity'))
                if not name:
                    skipped.append('忽略无名商品')
                    continue
                if qty is None or qty <= 0:
                    skipped.append(f'忽略数量无效的商品: {name}')
                    continue
                sku = _clean_cell(item.get('sku')) or _unique_sku(name[:6].upper())
                # 处理分类
                cat_id = _auto_category(name)
                # 处理供应商
                sup_id = None
                sup_name = _clean_cell(item.get('supplier'))
                if sup_name:
                    sups = SupplierModel.get_all()
                    smap = {s['name']: s['id'] for s in sups}
                    if sup_name in smap:
                        sup_id = smap[sup_name]
                    else:
                        sup_id = SupplierModel.create(sup_name)
                # 创建或查找商品
                prod = ProductModel.get_by_sku(sku) if sku else None
                if prod:
                    pid = prod['id']
                else:
                    pid = ProductModel.create(name, sku, cat_id, sup_id,
                        _clean_cell(item.get('unit')) or '个',
                        _clean_cell(item.get('specification')))
                # 入库
                up = _to_float(item.get('unit_price'))
                InventoryModel.stock_in(pid, qty, batch_no, _op('AI'), '入库单识别导入',
                                        unit_price=up, supplier_id=sup_id)
                imported.append({'name': name, 'sku': sku, 'quantity': qty, 'unit_price': up})

        return jsonify({'success': True, 'data': {
            'items': items, 'imported': imported, 'skipped': skipped, 'count': len(imported)}})
    except RuntimeError as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    except json.JSONDecodeError:
        return jsonify({'success': False, 'error': 'AI 识别结果解析失败，请重试或使用更清晰的图片'}), 400
    except Exception as e:
        _logger.error(f"入库单识别错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  页面路由
# ==========================================
@app.route('/')
def index():
    """主页"""
    return render_template('index.html')


@app.route('/static/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)


@app.route('/uploads/<path:filename>')
def download_upload(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


# ==========================================
#  仪表盘 API
# ==========================================
@app.route('/api/dashboard')
def api_dashboard():
    """获取仪表盘数据"""
    try:
        stats = StatsModel.get_dashboard()
        return jsonify({'success': True, 'data': stats})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  分类 API
# ==========================================
@app.route('/api/categories')
def api_categories():
    """获取所有分类"""
    try:
        categories = CategoryModel.get_all()
        return jsonify({'success': True, 'data': categories})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/categories', methods=['POST'])
def api_category_create():
    """创建分类"""
    try:
        data = request.json
        cat_id = CategoryModel.create(data['name'], data.get('description', ''))
        return jsonify({'success': True, 'id': cat_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/categories/<int:cat_id>', methods=['PUT'])
def api_category_update(cat_id):
    """更新分类"""
    try:
        data = request.json
        CategoryModel.update(cat_id, data['name'], data.get('description', ''))
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/categories/<int:cat_id>', methods=['DELETE'])
def api_category_delete(cat_id):
    """删除分类"""
    try:
        CategoryModel.delete(cat_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ==========================================
#  供应商 API
# ==========================================
@app.route('/api/suppliers')
def api_suppliers():
    try:
        suppliers = SupplierModel.get_all()
        return jsonify({'success': True, 'data': suppliers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/suppliers', methods=['POST'])
def api_supplier_create():
    try:
        d = request.json
        with db.transaction():
            sid = SupplierModel.create(d['name'], d.get('contact_person', ''), d.get('phone', ''),
                                        d.get('email', ''), d.get('address', ''), d.get('notes', ''))
            AuditLog.log('create', 'suppliers', sid, new_data=d, operator=_op())
        return jsonify({'success': True, 'id': sid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/suppliers/<int:sid>', methods=['PUT'])
def api_supplier_update(sid):
    try:
        d = request.json
        with db.transaction():
            old = SupplierModel.get_by_id(sid)
            SupplierModel.update(sid, d['name'], d.get('contact_person', ''), d.get('phone', ''),
                                 d.get('email', ''), d.get('address', ''), d.get('notes', ''))
            AuditLog.log('update', 'suppliers', sid, old_data=old, new_data=d, operator=_op())
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/suppliers/<int:sid>', methods=['DELETE'])
def api_supplier_delete(sid):
    try:
        with db.transaction():
            old = SupplierModel.get_by_id(sid)
            SupplierModel.delete(sid)
            AuditLog.log('delete', 'suppliers', sid, old_data=old, operator=_op())
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ==========================================
#  客户 API
# ==========================================
@app.route('/api/customers')
def api_customers():
    try:
        customers = CustomerModel.get_all()
        return jsonify({'success': True, 'data': customers})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/customers', methods=['POST'])
def api_customer_create():
    try:
        d = request.json
        with db.transaction():
            cid = CustomerModel.create(d['name'], d.get('contact_person', ''), d.get('phone', ''),
                                        d.get('email', ''), d.get('address', ''), d.get('notes', ''))
            AuditLog.log('create', 'customers', cid, new_data=d, operator=_op())
        return jsonify({'success': True, 'id': cid})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/customers/<int:cid>', methods=['PUT'])
def api_customer_update(cid):
    try:
        d = request.json
        with db.transaction():
            old = CustomerModel.get_by_id(cid)
            CustomerModel.update(cid, d['name'], d.get('contact_person', ''), d.get('phone', ''),
                                 d.get('email', ''), d.get('address', ''), d.get('notes', ''))
            AuditLog.log('update', 'customers', cid, old_data=old, new_data=d, operator=_op())
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/customers/<int:cid>', methods=['DELETE'])
def api_customer_delete(cid):
    try:
        with db.transaction():
            old = CustomerModel.get_by_id(cid)
            CustomerModel.delete(cid)
            AuditLog.log('delete', 'customers', cid, old_data=old, operator=_op())
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ==========================================
#  商品 API
# ==========================================
@app.route('/api/products')
def api_products():
    """获取所有商品"""
    try:
        keyword = request.args.get('search', '')
        if keyword:
            products = ProductModel.search(keyword)
        else:
            products = ProductModel.get_all()
        return jsonify({'success': True, 'data': products})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/<int:prod_id>')
def api_product_detail(prod_id):
    """获取商品详情"""
    try:
        product = ProductModel.get_by_id(prod_id)
        if not product:
            return jsonify({'success': False, 'error': '商品不存在'}), 404
        return jsonify({'success': True, 'data': product})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products', methods=['POST'])
def api_product_create():
    """创建商品"""
    try:
        data = request.json
        # 先校验数量，避免负数时「商品已创建落库却返回 400」的半截写
        qty = _to_int(data.get('quantity'))
        if qty is not None and qty < 0:
            return jsonify({'success': False, 'error': '数量不能为负数'}), 400
        with db.transaction():
            prod_id = ProductModel.create(
                name=data['name'],
                sku=data['sku'],
                category_id=data.get('category_id'),
                supplier_id=data.get('supplier_id'),
                unit=data.get('unit', '个'),
                specification=data.get('specification', ''),
                description=data.get('description', ''),
                unit_price=float(data.get('unit_price', 0) or 0),
                sale_price=float(data.get('sale_price', 0) or 0),
            )
            # 初始库存走 stock_in（生成入库流水），数量缺失/为 0 时跳过
            if qty is not None and qty > 0:
                InventoryModel.stock_in(
                    prod_id,
                    quantity=qty,
                    operator='初始创建',
                    notes='新建商品初始库存',
                    unit_price=float(data.get('unit_price', 0) or 0),
                    supplier_id=data.get('supplier_id'),
                )
            # 库位/阈值无条件写入（ProductModel.create 已保证库存行存在），
            # 修复「不填初始库存时库位/最低/最高库存全部丢失」的问题
            db.execute(
                "UPDATE inventory SET location=%s, min_stock=%s, max_stock=%s WHERE product_id=%s",
                (_clean_cell(data.get('location')),
                 _to_int(data.get('min_stock')) or 0,
                 _to_int(data.get('max_stock')) or 9999,
                 prod_id)
            )
            AuditLog.log('create', 'products', prod_id, new_data=data, operator=_op())
        return jsonify({'success': True, 'id': prod_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/products/<int:prod_id>', methods=['PUT'])
def api_product_update(prod_id):
    """更新商品（含库存信息）"""
    try:
        data = request.json
        # 先校验数量，避免负数时「商品字段已更新落库却返回 400」的半截写
        new_qty = _to_int(data.get('quantity'))
        if new_qty is not None and new_qty < 0:
            return jsonify({'success': False, 'error': '数量不能为负数'}), 400
        with db.transaction():
            old = ProductModel.get_by_id(prod_id)
            ProductModel.update(
                prod_id,
                name=data['name'],
                sku=data['sku'],
                category_id=data.get('category_id'),
                supplier_id=data.get('supplier_id'),
                unit=data.get('unit', '个'),
                specification=data.get('specification', ''),
                description=data.get('description', ''),
                unit_price=float(data.get('unit_price', 0) or 0),
                sale_price=float(data.get('sale_price', 0) or 0),
            )
            AuditLog.log('update', 'products', prod_id, old_data=old, new_data=data, operator=_op())
            # 同步更新库存信息（数量变化走出入库流水，保留审计）
            if new_qty is not None:
                _apply_inventory_change(
                    prod_id, new_qty,
                    location=_clean_cell(data.get('location')),
                    min_stock=_to_int(data.get('min_stock')),
                    max_stock=_to_int(data.get('max_stock')),
                )
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/products/<int:prod_id>', methods=['DELETE'])
def api_product_delete(prod_id):
    """删除商品"""
    try:
        with db.transaction():
            old = ProductModel.get_by_id(prod_id)
            ProductModel.delete(prod_id)
            AuditLog.log('delete', 'products', prod_id, old_data=old, operator=_op())
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ==========================================
#  库存 API
# ==========================================
@app.route('/api/inventory')
def api_inventory():
    """获取所有库存"""
    try:
        inv = InventoryModel.get_all()
        return jsonify({'success': True, 'data': inv})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/inventory/low-stock')
def api_low_stock():
    """获取低库存预警"""
    try:
        threshold = int(request.args.get('threshold', 10))
        items = InventoryModel.get_low_stock(threshold)
        return jsonify({'success': True, 'data': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/inventory/<int:product_id>', methods=['PUT'])
def api_inventory_update(product_id):
    """更新库存（数量变化走出入库流水，保留审计）"""
    try:
        data = request.json
        new_qty = _to_int(data.get('quantity'))
        if new_qty is None:
            return jsonify({'success': False, 'error': '缺少 quantity 参数'}), 400
        if new_qty < 0:
            return jsonify({'success': False, 'error': '数量不能为负数'}), 400
        with db.transaction():
            _apply_inventory_change(
                product_id, new_qty,
                location=_clean_cell(data.get('location')),
                min_stock=_to_int(data.get('min_stock')),
                max_stock=_to_int(data.get('max_stock')),
            )
        return jsonify({'success': True})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/inventory/stock-in', methods=['POST'])
def api_stock_in():
    """入库"""
    try:
        data = request.json
        up = float(data.get('unit_price', 0) or 0)
        sid = data.get('supplier_id')
        pid = int(data['product_id'])
        with db.transaction():
            InventoryModel.stock_in(
                product_id=pid,
                quantity=int(data['quantity']),
                batch_no=data.get('batch_no', ''),
                operator=_op(),
                notes=data.get('notes', ''),
                unit_price=up,
                supplier_id=sid,
            )
            # 更新商品最近进价和供应商
            if up > 0:
                db.execute("UPDATE products SET unit_price=%s WHERE id=%s", (up, pid))
            if sid:
                db.execute("UPDATE products SET supplier_id=%s WHERE id=%s", (sid, pid))
            AuditLog.log('stock_in', 'inventory', pid,
                         new_data={'qty': int(data['quantity']), 'unit_price': up, 'supplier_id': sid},
                         operator=_op())
        return jsonify({'success': True, 'message': '入库成功'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/inventory/stock-out', methods=['POST'])
def api_stock_out():
    """出库"""
    try:
        data = request.json
        with db.transaction():
            InventoryModel.stock_out(
                product_id=int(data['product_id']),
                quantity=int(data['quantity']),
                batch_no=data.get('batch_no', ''),
                operator=_op(),
                notes=data.get('notes', ''),
                customer_id=data.get('customer_id'),
                unit_price=float(data.get('unit_price', 0) or 0),
            )
            AuditLog.log('stock_out', 'inventory', int(data['product_id']),
                         new_data={'qty': int(data['quantity']), 'customer_id': data.get('customer_id')},
                         operator=_op())
        return jsonify({'success': True, 'message': '出库成功'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/order/submit', methods=['POST'])
def api_order_submit():
    """批量出库单（原子）：预检全部商品 → 单事务内逐项扣减 → 生成 Excel。

    修复 H5：原手工出库单前端逐项调 stock-out，第 2 项失败时第 1 项已扣减且无法回滚，
    造成「部分出库 + 提示出库成功」的半截写。本接口与 AI create_order 同语义：
    任一商品不存在/库存不足时整体不执行，不留半截数据。
    """
    try:
        data = request.json or {}
        items = data.get('items') or []
        customer_id = data.get('customer_id')
        operator = _op()   # 操作者 = 登录账号（Excel 经办人由 /api/order/export 单独接收）
        if not isinstance(items, list) or not items:
            return jsonify({'success': False, 'error': '请至少添加一个商品'}), 400
        if not customer_id:
            return jsonify({'success': False, 'error': '请选择客户'}), 400

        # 第一遍预检：校验商品存在与库存充足，不产生任何写入
        plan = []
        problems = []
        order_batch = _ai_batch_no('出库单')   # 每单独立批次，便于追溯
        for it in items:
            pid = _to_int(it.get('product_id'))
            qty = _to_int(it.get('quantity'))
            price = _to_float(it.get('unit_price'))
            prod = ProductModel.get_by_id(pid) if pid else None
            if not prod:
                problems.append(f"商品不存在: {pid}")
                continue
            if qty is None or qty <= 0:
                problems.append(f"数量无效: {prod['name']}")
                continue
            inv = db.query_one(
                "SELECT quantity FROM inventory WHERE product_id=%s FOR UPDATE", (pid,))
            avail = inv['quantity'] if inv else 0
            if avail < qty:
                problems.append(f"库存不足: {prod['name']} 当前{avail} 需{qty}")
                continue
            plan.append({'prod': prod, 'qty': qty, 'price': price or 0})

        if problems:
            return jsonify({'success': False, 'error': '出库单未执行：' + '；'.join(problems)}), 400

        # 全部通过 → 单事务内逐项原子扣减（FOR UPDATE 已锁行，杜绝并发超卖），失败整体回滚
        done = []
        with db.transaction():
            for p in plan:
                InventoryModel.stock_out(p['prod']['id'], p['qty'], order_batch, operator,
                                         customer_id=int(customer_id), unit_price=p['price'])
                done.append({'product_id': p['prod']['id'], 'name': p['prod']['name'],
                             'sku': p['prod']['sku'], 'quantity': p['qty'],
                             'unit': p['prod'].get('unit', ''),
                             'specification': p['prod'].get('specification', ''),
                             'sale_price': p['price'],
                             'batch_no': order_batch,
                             'unit_price': p['price']})
            # record_id 为整数列，批次号字符串无法写入 → 用 None，批次号放入 new_data
            AuditLog.log('stock_out', 'inventory', None,
                         new_data={'batch_no': order_batch, 'count': len(done)},
                         operator=operator)

        return jsonify({'success': True, 'data': {'batch_no': order_batch, 'count': len(done), 'items': done}})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        _logger.error(f"批量出库单失败: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  交易记录 API
# ==========================================
@app.route('/api/transactions')
def api_transactions():
    """获取交易记录"""
    try:
        limit = int(request.args.get('limit', 100))
        txn = TransactionModel.get_all(limit)
        return jsonify({'success': True, 'data': txn})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/transactions/stats')
def api_transaction_stats():
    """获取交易统计"""
    try:
        days = int(request.args.get('days', 30))
        stats = TransactionModel.get_stats(days)
        return jsonify({'success': True, 'data': stats})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  Excel 上传 API
# ==========================================
@app.route('/api/upload', methods=['POST'])
def api_upload():
    """上传 Excel 文件并导入数据"""
    filepath = None
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '请选择文件'}), 400

        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': '仅支持 .xlsx .xls .csv 格式'}), 400

        # 导入模式：replace=全量覆盖, increment=增量累加
        mode = request.form.get('mode', 'replace')

        original_name = file.filename
        ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
        # 存储用 uuid 命名（secure_filename 会剥掉中文名），原始文件名仅入 DB 记录
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        file_size = os.path.getsize(filepath)

        # 记录上传
        upload_id = ExcelUploadModel.create(original_name, file_size)
        ExcelUploadModel.update_status(upload_id, 'processing')

        # 解析 Excel / CSV
        try:
            if ext == 'csv':
                try:
                    df = pd.read_csv(filepath, encoding='utf-8-sig')
                except UnicodeDecodeError:
                    df = pd.read_csv(filepath, encoding='gbk')
            elif ext == 'xls':
                df = pd.read_excel(filepath, engine='xlrd')
            else:
                df = pd.read_excel(filepath, engine='openpyxl')

            # 映射列名（支持中英文列名）
            col_map = {
                '商品名称': 'name', '名称': 'name', 'name': 'name',
                'SKU': 'sku', 'sku': 'sku', '编码': 'sku', '编号': 'sku',
                '分类': 'category', 'category': 'category',
                '单位': 'unit', 'unit': 'unit',
                '规格': 'specification', 'specification': 'specification',
                '数量': 'quantity', 'quantity': 'quantity', '库存': 'quantity',
                '位置': 'location', 'location': 'location', '库位': 'location',
                '最低库存': 'min_stock', 'min_stock': 'min_stock',
                '最高库存': 'max_stock', 'max_stock': 'max_stock',
                '描述': 'description', 'description': 'description',
            }

            # 标准化列名
            df.columns = [str(c).strip() for c in df.columns]
            rename_map = {}
            for col in df.columns:
                if col in col_map:
                    rename_map[col] = col_map[col]
            df.rename(columns=rename_map, inplace=True)

            rows_imported = 0
            errors = []
            seen_skus = set()   # 文件内重复 SKU 检测
            loc_counters = {}

            # 多行写入包在一个事务里：意外异常整体回滚，不留半截数据
            with db.transaction():
                for idx, row in df.iterrows():
                    try:
                        name = _clean_cell(row.get('name'))
                        sku = _clean_cell(row.get('sku'))

                        if not name:
                            continue
                        if sku and sku in seen_skus:
                            errors.append(f"第 {idx+2} 行: 文件内 SKU 重复({sku})，已跳过")
                            continue
                        if sku:
                            seen_skus.add(sku)

                        # 处理分类
                        cat_name = _clean_cell(row.get('category'))
                        category_id = None
                        if cat_name:
                            existing_cat = CategoryModel.get_all()
                            cat_map = {c['name']: c['id'] for c in existing_cat}
                            if cat_name in cat_map:
                                category_id = cat_map[cat_name]
                            else:
                                category_id = CategoryModel.create(cat_name)

                        # 自动生成 SKU
                        if not sku:
                            prefix = (cat_name[:3] if cat_name else name[:3]).upper()
                            counter = 1
                            while True:
                                candidate = f'{prefix}-{counter:04d}'
                                if not ProductModel.get_by_sku(candidate):
                                    break
                                counter += 1
                            sku = candidate

                        # 创建或更新商品（空单元格统一清洗，避免写入字面 "nan"）
                        unit = _clean_cell(row.get('unit')) or '个'
                        specification = _clean_cell(row.get('specification'))
                        description = _clean_cell(row.get('description'))
                        existing = ProductModel.get_by_sku(sku)
                        if existing:
                            # 文件没有价格/供应商列 → 保留原值，避免重新导入时把价格清零、供应商抹掉；
                            # 分类仅在本行显式提供时才覆盖
                            ProductModel.update(
                                existing['id'], name=name, sku=sku,
                                category_id=category_id if category_id is not None else existing.get('category_id'),
                                supplier_id=existing.get('supplier_id'),
                                unit=unit, specification=specification, description=description,
                            )
                            prod_id = existing['id']
                        else:
                            prod_id = ProductModel.create(
                                name=name, sku=sku, category_id=category_id,
                                unit=unit, specification=specification, description=description,
                            )

                        # 更新库存（对比差异，自动生成出入库记录）
                        qty_raw = row.get('quantity')
                        qty = _to_int(qty_raw) if (qty_raw is not None and _clean_cell(qty_raw) != '') else None
                        if qty is not None:
                            if qty < 0:
                                errors.append(f"第 {idx+2} 行: 数量为负({qty})，已跳过库存更新")
                            else:
                                # 查询原库存（用于差异计算、流水记录，以及文件缺列时保留原有库位/阈值）
                                # FOR UPDATE 行锁：Excel 全量覆盖模式下 diff=目标-当前 的读取必须锁定，
                                # 否则两个导入并发时会按同一旧值算 diff，最后落库值 ≠ 文件目标值。
                                old_inv = db.query_one(
                                    "SELECT quantity, location, min_stock, max_stock FROM inventory WHERE product_id = %s FOR UPDATE",
                                    (prod_id,)
                                )
                                old_qty = old_inv['quantity'] if old_inv else 0

                                # 库位：文件没有库位列时保留原库位，避免重导入把已有库位冲掉
                                location = _clean_cell(row.get('location'))
                                if not location and old_inv and _clean_cell(old_inv.get('location')):
                                    location = old_inv['location']

                                # 自动生成库位（仅当没有可保留的原有库位时）
                                if not location:
                                    cat_key = cat_name[:2] if cat_name else 'ZZ'
                                    if cat_key not in loc_counters:
                                        # 未分类商品 category_id=None → 必须用 IS NULL，
                                        # 否则 WHERE p.category_id=%s 永远查不到、每次上传都从 ZZ-001 重来
                                        if category_id is None:
                                            loc_rows = db.query(
                                                """SELECT i.location FROM inventory i
                                                   JOIN products p ON i.product_id=p.id
                                                   WHERE p.category_id IS NULL AND i.location LIKE %s""",
                                                (f'{cat_key}-%',))
                                        else:
                                            loc_rows = db.query(
                                                """SELECT i.location FROM inventory i
                                                   JOIN products p ON i.product_id=p.id
                                                   WHERE p.category_id=%s AND i.location LIKE %s""",
                                                (category_id, f'{cat_key}-%'))
                                        # 取数值最大（字符串排序在超过 999 个库位时会把 'ZZ-1000' 排在 'ZZ-999' 前）
                                        nums = []
                                        for r in loc_rows:
                                            try:
                                                nums.append(int(r['location'].split('-')[1]))
                                            except (ValueError, IndexError, TypeError):
                                                continue
                                        loc_counters[cat_key] = max(nums) + 1 if nums else 1
                                    else:
                                        loc_counters[cat_key] += 1
                                    location = f'{cat_key}-{loc_counters[cat_key]:03d}'

                                # 阈值：文件显式提供才更新，否则保留原值（重导入无阈值列的文件不再抹掉低库存预警）
                                min_raw = row.get('min_stock')
                                max_raw = row.get('max_stock')
                                if min_raw is not None and _clean_cell(min_raw) != '':
                                    min_stock = _to_int(min_raw) or 0
                                else:
                                    min_stock = old_inv['min_stock'] if old_inv else 0
                                if max_raw is not None and _clean_cell(max_raw) != '':
                                    max_stock = _to_int(max_raw) or 9999
                                else:
                                    max_stock = old_inv['max_stock'] if old_inv else 9999

                                # 增量模式：Excel 数量为新增量，累加到现有库存
                                diff = qty if mode == 'increment' else qty - old_qty

                                # 原子增减，避免并发「读-改-写」丢失更新
                                skipped_stock = False
                                if diff > 0:
                                    affected, _ = db.execute(
                                        "UPDATE inventory SET quantity = quantity + %s WHERE product_id=%s",
                                        (diff, prod_id))
                                    if affected == 0:
                                        # 库存行缺失（历史数据）→ 补建后再自增
                                        try:
                                            db.execute("INSERT INTO inventory (product_id, quantity) VALUES (%s, %s)", (prod_id, diff))
                                        except pymysql.err.IntegrityError:
                                            db.execute(
                                                "UPDATE inventory SET quantity = quantity + %s WHERE product_id=%s",
                                                (diff, prod_id))
                                elif diff < 0:
                                    dec = -diff
                                    affected, _ = db.execute(
                                        "UPDATE inventory SET quantity = quantity - %s WHERE product_id=%s AND quantity >= %s",
                                        (dec, prod_id, dec))
                                    if affected == 0:
                                        errors.append(f"第 {idx+2} 行: 库存不足（当前 {old_qty}，需减 {dec}），已跳过")
                                        skipped_stock = True

                                # 库位/阈值与数量变更相互独立，单独更新
                                db.execute(
                                    "UPDATE inventory SET location=%s, min_stock=%s, max_stock=%s WHERE product_id=%s",
                                    (location, min_stock, max_stock, prod_id))

                                # 如果数量有变化，生成出入库记录（before/after 以实际落库值为准）
                                if not skipped_stock and diff != 0:
                                    after_inv = db.query_one(
                                        "SELECT quantity FROM inventory WHERE product_id=%s", (prod_id,))
                                    new_qty = after_inv['quantity'] if after_inv else old_qty + diff
                                    before_qty = new_qty - diff
                                    txn_type = 'in' if diff > 0 else 'out'
                                    batch_no = f'Excel-{upload_id}'
                                    db.execute(
                                        """INSERT INTO transactions
                                           (product_id, type, quantity, before_qty, after_qty, batch_no, operator, notes)
                                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                                        (prod_id, txn_type, abs(diff), before_qty, new_qty,
                                         batch_no, _op('Excel导入'), f'文件: {original_name}'))

                        rows_imported += 1
                    except Exception as e:
                        errors.append(f"第 {idx+2} 行: {str(e)}")

            err_summary = f"{len(errors)} 行出错" if errors else ''
            # 有行级错误时标记 partial，不再一律 success（上传记录里能看出部分失败）
            status = 'partial' if errors else 'success'
            ExcelUploadModel.update_status(upload_id, status, rows_imported, err_summary)

            return jsonify({
                'success': True,
                'data': {
                    'upload_id': upload_id,
                    'status': status,
                    'rows_imported': rows_imported,
                    'total_rows': len(df),
                    'errors': errors[:20],  # 最多返回20条错误
                }
            })

        except Exception as e:
            ExcelUploadModel.update_status(upload_id, 'failed', 0, str(e))
            return jsonify({'success': False, 'error': f'文件解析失败: {str(e)}'}), 400

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        # 清理临时上传文件，避免 uploads/ 无限增长
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass


@app.route('/api/uploads')
def api_uploads():
    """获取上传记录"""
    try:
        records = ExcelUploadModel.get_recent()
        return jsonify({'success': True, 'data': records})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==========================================
#  供应商/客户 Excel 导入
# ==========================================
@app.route('/api/upload/suppliers', methods=['POST'])
def api_upload_suppliers():
    filepath = None
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择文件'}), 400
        file = request.files['file']
        original_name = file.filename
        ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({'success': False, 'error': '仅支持 .xlsx .xls .csv 格式'}), 400
        filename = f"supplier_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        df = _read_upload_df(filepath, ext)

        col_map = {'名称': 'name', 'name': 'name', '供应商名称': 'name',
                   '联系人': 'contact', '电话': 'phone', 'phone': 'phone',
                   '邮箱': 'email', 'email': 'email',
                   '地址': 'address', 'address': 'address',
                   '备注': 'notes', 'notes': 'notes'}
        df.columns = [str(c).strip() for c in df.columns]
        rename_map = {c: col_map[c] for c in df.columns if c in col_map}
        df.rename(columns=rename_map, inplace=True)

        created = 0
        updated = 0
        with db.transaction():
            for _, row in df.iterrows():
                name = _clean_cell(row.get('name'))
                if not name:
                    continue
                contact = _clean_cell(row.get('contact'))
                phone = _clean_cell(row.get('phone'))
                email = _clean_cell(row.get('email'))
                address = _clean_cell(row.get('address'))
                notes = _clean_cell(row.get('notes'))
                existing = SupplierModel.get_by_name(name)
                if existing:
                    # 按名称去重：重复上传更新原记录（空单元格保留原值），不再产生重复供应商
                    SupplierModel.update(
                        existing['id'], name,
                        contact or (existing.get('contact_person') or ''),
                        phone or (existing.get('phone') or ''),
                        email or (existing.get('email') or ''),
                        address or (existing.get('address') or ''),
                        notes or (existing.get('notes') or ''))
                    updated += 1
                else:
                    SupplierModel.create(name, contact, phone, email, address, notes)
                    created += 1
        return jsonify({'success': True, 'data': {'rows_imported': created + updated,
                                                  'created': created, 'updated': updated}})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass


@app.route('/api/upload/customers', methods=['POST'])
def api_upload_customers():
    filepath = None
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '请选择文件'}), 400
        file = request.files['file']
        original_name = file.filename
        ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
        if ext not in ALLOWED_EXTENSIONS:
            return jsonify({'success': False, 'error': '仅支持 .xlsx .xls .csv 格式'}), 400
        filename = f"customer_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        df = _read_upload_df(filepath, ext)

        col_map = {'名称': 'name', 'name': 'name', '客户名称': 'name',
                   '联系人': 'contact', '电话': 'phone', 'phone': 'phone',
                   '邮箱': 'email', 'email': 'email',
                   '地址': 'address', 'address': 'address',
                   '备注': 'notes', 'notes': 'notes'}
        df.columns = [str(c).strip() for c in df.columns]
        rename_map = {c: col_map[c] for c in df.columns if c in col_map}
        df.rename(columns=rename_map, inplace=True)

        created = 0
        updated = 0
        with db.transaction():
            for _, row in df.iterrows():
                name = _clean_cell(row.get('name'))
                if not name:
                    continue
                contact = _clean_cell(row.get('contact'))
                phone = _clean_cell(row.get('phone'))
                email = _clean_cell(row.get('email'))
                address = _clean_cell(row.get('address'))
                notes = _clean_cell(row.get('notes'))
                existing = CustomerModel.get_by_name(name)
                if existing:
                    # 按名称去重：重复上传更新原记录（空单元格保留原值），不再产生重复客户
                    CustomerModel.update(
                        existing['id'], name,
                        contact or (existing.get('contact_person') or ''),
                        phone or (existing.get('phone') or ''),
                        email or (existing.get('email') or ''),
                        address or (existing.get('address') or ''),
                        notes or (existing.get('notes') or ''))
                    updated += 1
                else:
                    CustomerModel.create(name, contact, phone, email, address, notes)
                    created += 1
        return jsonify({'success': True, 'data': {'rows_imported': created + updated,
                                                  'created': created, 'updated': updated}})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    finally:
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass


# ==========================================
#  AI 分析 API
# ==========================================
@app.route('/api/ai/health')
def api_ai_health():
    """AI 服务健康检查"""
    ok, msg = ai_service.health_check()
    return jsonify({'success': ok, 'message': msg})


@app.route('/api/ai/analyze')
def api_ai_analyze():
    """AI 库存分析"""
    try:
        query_type = request.args.get('type', 'general')
        result = ai_service.analyze_inventory(query_type)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        _logger.error(f"AI 分析错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ai/smart-import', methods=['POST'])
def api_ai_smart_import():
    """AI 智能导入：自然语言 → 解析 → 入库"""
    try:
        data = request.json
        text = data.get('text', '')
        if not text:
            return jsonify({'success': False, 'error': '请输入采购信息'}), 400

        # 1. AI 解析自然语言
        items = ai_service.smart_import(text)
        if not isinstance(items, list):
            return jsonify({'success': False, 'error': 'AI 返回格式异常，请重新描述'}), 400

        # 2. 逐个导入（整体事务，异常回滚，避免半截数据）
        imported = []
        # 本次导入独立成批（旧常量 'AI智能导入' 会让多次导入在批次视图里合并）
        batch_no = _ai_batch_no('AI导入')
        with db.transaction():
            for item in items:
                if not isinstance(item, dict):
                    continue
                name = _clean_cell(item.get('name'))
                sku = _clean_cell(item.get('sku'))
                category_name = _clean_cell(item.get('category'))
                unit = _clean_cell(item.get('unit')) or '个'
                qty = _to_int(item.get('quantity'))
                notes = _clean_cell(item.get('notes'))

                if not name or not sku or qty is None or qty <= 0:
                    continue

                # 处理分类
                category_id = None
                if category_name:
                    existing_cats = CategoryModel.get_all()
                    cat_map = {c['name']: c['id'] for c in existing_cats}
                    if category_name in cat_map:
                        category_id = cat_map[category_name]
                    else:
                        category_id = CategoryModel.create(category_name)

                # 处理供应商
                supplier_id = None
                supplier_name = _clean_cell(item.get('supplier') or item.get('supplier_name'))
                if supplier_name:
                    existing_sups = SupplierModel.get_all()
                    sup_map = {s['name']: s['id'] for s in existing_sups}
                    if supplier_name in sup_map:
                        supplier_id = sup_map[supplier_name]
                    else:
                        supplier_id = SupplierModel.create(supplier_name)

                # 查找或创建商品（已有商品：解析结果无价格列 → 保留原价格；分类/供应商仅在显式提供时覆盖）
                existing = ProductModel.get_by_sku(sku)
                if existing:
                    prod_id = existing['id']
                    ProductModel.update(prod_id, name, sku,
                                        category_id or existing.get('category_id'),
                                        supplier_id or existing.get('supplier_id'), unit,
                                        _clean_cell(item.get('specification')), notes)
                else:
                    prod_id = ProductModel.create(name, sku, category_id, supplier_id, unit,
                                                  _clean_cell(item.get('specification')), notes)

                # 入库（原子自增，避免并发「读-改-写」丢失更新）
                db.execute(
                    """INSERT INTO inventory (product_id, quantity) VALUES (%s, %s)
                       ON DUPLICATE KEY UPDATE quantity = quantity + VALUES(quantity)""",
                    (prod_id, qty)
                )
                after_inv = db.query_one("SELECT quantity FROM inventory WHERE product_id=%s", (prod_id,))
                new_qty = after_inv['quantity'] if after_inv else qty
                old_qty = new_qty - qty
                item_price = _to_float(item.get('price'))
                db.execute(
                    """INSERT INTO transactions
                       (product_id, type, quantity, unit_price, supplier_id, before_qty, after_qty, batch_no, operator, notes)
                       VALUES (%s, 'in', %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (prod_id, qty, item_price, supplier_id, old_qty, new_qty, batch_no, _op('AI'), notes))
                imported.append({'name': name, 'sku': sku, 'quantity': qty})

        return jsonify({
            'success': True,
            'data': {
                'parsed_text': text,
                'items': items,
                'imported': imported,
                'count': len(imported),
            }
        })

    except json.JSONDecodeError as e:
        _logger.error(f"AI 解析失败: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'AI 解析结果格式错误，请重新描述: {str(e)}'}), 400
    except ValueError as e:
        _logger.error(f"AI 解析失败: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        _logger.error(f"智能导入错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ai/chat', methods=['POST'])
def api_ai_chat():
    """AI 对话（支持直接修改库存）"""
    try:
        data = request.json
        message = data.get('message', '')
        if not message:
            return jsonify({'success': False, 'error': '请输入问题'}), 400

        result = ai_service.chat(message)
        actions, reply = ai_service.parse_actions(result)

        # 执行 AI 指令（每个动作独立容错，失败不会中断整轮，也不会留下 500）
        log = []
        if actions:
            for act in actions:
                try:
                    if not isinstance(act, dict):
                        continue
                    action_type = act.get('action', '')
                    sku = _clean_cell(act.get('sku'))
                    qty = _to_int(act.get('quantity'))
                    notes = _clean_cell(act.get('notes'))

                    # 供应商/客户操作（不需 sku/quantity）；按名称去重，已存在则跳过不覆盖
                    if action_type == 'add_supplier':
                        name_s = _clean_cell(act.get('name'))
                        if name_s:
                            if SupplierModel.get_by_name(name_s):
                                log.append(f"供应商已存在，跳过: {name_s}")
                            else:
                                SupplierModel.create(name_s, act.get('contact', ''), act.get('phone', ''))
                                log.append(f"新增供应商: {name_s}")
                        continue
                    if action_type == 'add_customer':
                        name_c = _clean_cell(act.get('name'))
                        if name_c:
                            if CustomerModel.get_by_name(name_c):
                                log.append(f"客户已存在，跳过: {name_c}")
                            else:
                                CustomerModel.create(name_c, act.get('contact', ''), act.get('phone', ''))
                                log.append(f"新增客户: {name_c}")
                        continue

                    if action_type == 'create_order':
                        # AI 创建出库单：先预检全部商品，全部通过才在事务里建客户+扣减，杜绝部分出库与垃圾客户
                        cust_name = (act.get('customer') or '').strip()
                        op = _op('AI')
                        keeper = act.get('keeper') or ''
                        wh = act.get('warehouse') or ''
                        order_items = act.get('items') or []
                        if not isinstance(order_items, list):
                            order_items = []

                        # 预检（不产生任何写入）
                        plan = []
                        problems = []
                        for oi in order_items:
                            if not isinstance(oi, dict):
                                continue
                            sku_i = _clean_cell(oi.get('sku'))
                            qty_i = _to_int(oi.get('quantity'))
                            price_i = _to_float(oi.get('price'))
                            prod = ProductModel.get_by_sku(sku_i) if sku_i else None
                            if not prod:
                                problems.append(f"商品不存在: {sku_i or '(空)'}")
                                continue
                            if qty_i is None or qty_i <= 0:
                                problems.append(f"数量无效: {prod['name']}")
                                continue
                            inv = db.query_one("SELECT quantity FROM inventory WHERE product_id=%s", (prod['id'],))
                            avail = inv['quantity'] if inv else 0
                            if avail < qty_i:
                                problems.append(f"库存不足: {prod['name']} 当前{avail} 需{qty_i}")
                                continue
                            plan.append((prod, sku_i, qty_i, price_i))

                        if problems:
                            log.append("出库单未执行: " + "；".join(problems))
                            continue

                        # 全部通过 → 事务内建客户+扣减（预检失败时不会留下垃圾客户）
                        done = []
                        order_batch = _ai_batch_no('AI出库')   # 每单独立批次，便于追溯
                        with db.transaction():
                            cid = None
                            if cust_name:
                                existing_cust = CustomerModel.get_by_name(cust_name)
                                cid = existing_cust['id'] if existing_cust else CustomerModel.create(cust_name)
                            for prod, sku_i, qty_i, price_i in plan:
                                InventoryModel.stock_out(prod['id'], qty_i, order_batch, op,
                                                         customer_id=cid, unit_price=price_i)
                                done.append({'name': prod['name'], 'sku': sku_i, 'quantity': qty_i,
                                             'sale_price': price_i, 'specification': prod.get('specification', '')})
                        # 生成 Excel
                        import openpyxl
                        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                        from io import BytesIO
                        now_str = datetime.now().strftime('%Y-%m-%d')
                        wb = openpyxl.Workbook(); ws = wb.active; ws.title = '出库单'
                        tf = Font(size=16, bold=True); hf = Font(color='FFFFFF', bold=True, size=11)
                        hfill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
                        bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                        ca = Alignment(horizontal='center', vertical='center')
                        ra = Alignment(horizontal='right', vertical='center')
                        la = Alignment(horizontal='left', vertical='center')
                        ws.merge_cells('A1:G1'); ws['A1'] = '出 库 单'; ws['A1'].font = tf; ws['A1'].alignment = ca
                        ws.cell(row=2, column=1, value='日期:').alignment = ra
                        ws.cell(row=2, column=2, value=now_str).alignment = la
                        ws.cell(row=2, column=4, value='客户:').alignment = ra
                        ws.cell(row=2, column=5, value=cust_name).alignment = la
                        for ci, h in enumerate(['序号','商品名称','规格','SKU','数量','单价(元)','金额(元)'], 1):
                            c = ws.cell(row=3, column=ci, value=h); c.fill = hfill; c.font = hf; c.alignment = ca; c.border = bd
                        total = 0
                        for idx, oi in enumerate(done, 1):
                            r = 3 + idx; sub = round(oi['sale_price'] * oi['quantity'], 2); total += sub
                            for ci, v in enumerate([idx, oi['name'], oi['specification'], oi['sku'],
                                                     oi['quantity'], oi['sale_price'], sub], 1):
                                c = ws.cell(row=r, column=ci, value=v); c.border = bd
                                c.alignment = ca if ci < 6 else ra
                        sr = 3 + len(done) + 1
                        ws.merge_cells(f'A{sr}:E{sr}'); ws.cell(row=sr, column=1, value='合计').font = Font(bold=True)
                        ws.cell(row=sr, column=1).alignment = ra
                        tc = ws.cell(row=sr, column=7, value=total); tc.font = Font(bold=True, color='FF0000')
                        tc.border = bd; tc.alignment = ra
                        bot = sr + 1
                        ws.cell(row=bot, column=1, value='经办人:').alignment = ra
                        ws.cell(row=bot, column=2, value=op).alignment = la
                        ws.cell(row=bot, column=4, value='库管:').alignment = ra
                        ws.cell(row=bot, column=5, value=keeper).alignment = la
                        ws.cell(row=bot+1, column=1, value='仓库编号:').alignment = ra
                        ws.cell(row=bot+1, column=2, value=wh).alignment = la
                        for i, w in enumerate([6,22,20,15,8,12,12], 1):
                            ws.column_dimensions[chr(64+i)].width = w
                        safe_cust = _safe_filename_part(cust_name)
                        fname = f"出库单_{safe_cust}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        fpath = os.path.join(UPLOAD_FOLDER, fname)
                        wb.save(fpath)
                        _cleanup_old_order_files()   # 顺手清理过期出库单，防止 uploads/ 无限增长
                        log.append(f"出库单: {cust_name}, {len(done)}项, 合计¥{total}")
                        reply = (reply or '出库单已创建') + f'\n\n📥 [点击下载出库单](/uploads/{fname})'
                        continue

                    if action_type == 'smart_import':
                        items = ai_service.smart_import(act.get('text', ''))
                        if not isinstance(items, list):
                            log.append("智能导入: AI 返回格式异常")
                            continue
                        for item in items:
                            if not isinstance(item, dict):
                                continue
                            qty_imp = _to_int(item.get('quantity'))
                            if qty_imp is None or qty_imp <= 0:
                                continue
                            _do_ai_stock_in(_clean_cell(item.get('sku')), qty_imp,
                                            _clean_cell(item.get('name')), _clean_cell(item.get('notes')))
                            log.append(f"智能导入: {item.get('name')} +{qty_imp}")
                        continue

                    if not sku or qty is None or qty <= 0:
                        continue

                    if action_type == 'stock_in':
                        _do_ai_stock_in(sku, qty, act.get('name', ''),
                                        unit=act.get('unit', ''), category_name=act.get('category', ''),
                                        supplier_name=act.get('supplier', ''))
                        log.append(f"入库: {sku} +{qty}")
                    elif action_type == 'stock_out':
                        _do_ai_stock_out(sku, qty, notes)
                        log.append(f"出库: {sku} -{qty}")
                    elif action_type == 'set_quantity':
                        _do_ai_set_quantity(sku, qty, notes)
                        log.append(f"调库: {sku} = {qty}")
                    else:
                        log.append(f"未知操作: {action_type}")
                except Exception as e:
                    _logger.error(f"AI 动作执行失败: {traceback.format_exc()}")
                    log.append(f"操作执行失败: {str(e)}")

        reply = reply or result
        if log:
            reply += f'\n\n✅ 已执行: {"；".join(log)}'

        return jsonify({'success': True, 'data': reply, 'actions': log})
    except Exception as e:
        _logger.error(f"AI 对话错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


# 关键字 → 分类自动映射
CATEGORY_KEYWORDS = {
    '电子产品': ['电阻', '电容', '芯片', 'PCB', '二极管', 'LED', '传感器', '单片机', '电源', '电池', '模块', '电路板', '数据线', '连接线', '充电器', '适配器'],
    '机械零件': ['螺丝', '螺母', '垫圈', '轴承', '齿轮', '弹簧', '密封', '阀门', '法兰', '螺栓', '销', '轴', '链条', '皮带'],
    '原材料': ['钢材', '铝', '铜', '塑料', '橡胶', '玻璃', '木材', '布料', '皮革', '海绵', '泡棉'],
    '包装材料': ['纸箱', '胶带', '气泡', '标签', '打包', '泡沫', '缠绕', '封口', '包装袋', '快递袋'],
    '办公用品': ['打印纸', '笔', '文件夹', '订书', '胶水', '便签', '笔记本', '墨盒', '硒鼓', '档案'],
    '五金工具': ['扳手', '钳子', '锤子', '螺丝刀', '电钻', '锯', '尺', '刀具', '剪刀', '卷尺'],
}

def _auto_category(name):
    """根据商品名自动推断分类"""
    for cat, keywords in CATEGORY_KEYWORDS.items():
        for kw in keywords:
            if kw in name:
                cats = CategoryModel.get_all()
                cat_map = {c['name']: c['id'] for c in cats}
                if cat in cat_map:
                    return cat_map[cat]
                return CategoryModel.create(cat)
    return None


def _do_ai_stock_in(sku, qty, name_hint='', notes='', unit='', category_name='', supplier_name=''):
    """AI 执行入库（含自动分类和供应商关联），整体一个事务，失败不留半截数据"""
    prod = ProductModel.get_by_sku(sku)
    with db.transaction():
        if not prod:
            # 自动推断分类
            cat_id = _auto_category(name_hint or sku)
            if not cat_id and category_name:
                cats = CategoryModel.get_all()
                cat_map = {c['name']: c['id'] for c in cats}
                if category_name in cat_map:
                    cat_id = cat_map[category_name]
                else:
                    cat_id = CategoryModel.create(category_name)
            # 处理供应商
            sup_id = None
            if supplier_name:
                sups = SupplierModel.get_all()
                sup_map = {s['name']: s['id'] for s in sups}
                if supplier_name in sup_map:
                    sup_id = sup_map[supplier_name]
                else:
                    sup_id = SupplierModel.create(supplier_name)
            # 创建商品
            prod_id = ProductModel.create(
                name_hint or sku, sku, category_id=cat_id, supplier_id=sup_id,
                unit=unit or '个'
            )
        else:
            prod_id = prod['id']
            sup_id = prod.get('supplier_id')  # 已有商品用已有供应商
        InventoryModel.stock_in(prod_id, qty, _ai_batch_no('AI操作'), _op('AI'), notes, supplier_id=sup_id)
        # 更新供应商
        if sup_id:
            db.execute("UPDATE products SET supplier_id=%s WHERE id=%s", (sup_id, prod_id))


def _do_ai_stock_out(sku, qty, notes=''):
    """AI 执行出库（扣减+流水同一事务）"""
    prod = ProductModel.get_by_sku(sku)
    if not prod:
        raise ValueError(f'商品 {sku} 不存在')
    with db.transaction():
        InventoryModel.stock_out(prod['id'], qty, _ai_batch_no('AI操作'), _op('AI'), notes)


def _do_ai_set_quantity(sku, qty, notes=''):
    """AI 执行库存调整（原子增减+流水，避免并发「读-改-写」丢失更新）"""
    if qty < 0:
        raise ValueError('库存数量不能为负数')
    prod = ProductModel.get_by_sku(sku)
    if not prod:
        raise ValueError(f'商品 {sku} 不存在')
    with db.transaction():
        # FOR UPDATE 行锁：消除「设为绝对值」读取旧值后并发算 delta 的丢失更新
        old = db.query_one("SELECT quantity FROM inventory WHERE product_id=%s FOR UPDATE", (prod['id'],))
        old_qty = old['quantity'] if old else 0
        delta = qty - old_qty
        note = notes or f'AI调整库存至 {qty}'
        batch_no = _ai_batch_no('AI操作')   # 本次调整独立成批（增/减共用同一批次）
        if delta > 0:
            InventoryModel.stock_in(prod['id'], delta, batch_no, _op('AI'), note)
        elif delta < 0:
            # 库存不足时 stock_out 抛错，事务回滚，不会出现负库存
            InventoryModel.stock_out(prod['id'], -delta, batch_no, _op('AI'), note)


# ==========================================
#  启动
# ==========================================
if __name__ == '__main__':
    print("=" * 60)
    print("  仓库管理系统启动中...")
    print("  访问地址: http://0.0.0.0:5050")
    print("  AI 服务: http://100.101.108.100:1234")
    print("=" * 60)
    app.run(host='0.0.0.0', port=5050, debug=True)
