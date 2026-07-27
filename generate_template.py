"""
生成 Excel 导入模板，含商品、供应商、客户三个工作表
运行: python generate_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()

header_style = {
    'fill': PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid'),
    'font': Font(color='FFFFFF', bold=True, size=11),
    'alignment': Alignment(horizontal='center'),
}

data_align = Alignment(horizontal='center')

def write_sheet(ws, title, headers, data, col_widths):
    ws.title = title
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.alignment = header_style['alignment']
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width


# ==========================================
#  Sheet 1: 商品导入
# ==========================================
ws1 = wb.active
product_headers = ['商品名称*', 'SKU', '分类', '供应商', '单位', '规格', '数量', '进货价', '售价', '库位', '最低库存', '最高库存', '描述']
product_data = [
    ['螺丝 M6x20', 'SCR-M6-020', '机械零件', '华强五金', '盒', '不锈钢 M6x20mm', 1500, 0.05, 0.15, 'A-01-03', 200, 5000, ''],
    ['电阻 10KΩ', 'RES-10K', '电子产品', '深华电子', '个', '0805 贴片 ±5%', 300, 0.01, 0.05, 'B-02-01', 500, 10000, ''],
    ['打印纸 A4', 'PAP-A4', '办公用品', '', '包', '500张/包 80g', 80, 0, 20, 'C-01-05', 50, 500, ''],
    ['电容 100μF', 'CAP-100U', '电子产品', '深华电子', '个', '铝电解 25V', 1200, 0.03, 0.1, 'B-01-04', 300, 5000, ''],
    ['打包胶带', 'TAP-001', '包装材料', '', '卷', '48mm×50m', 120, 0, 3, 'C-02-02', 40, 500, ''],
    ['佳能R8相机', '', '摄影器材', '佳能', '台', '全画幅微单', 2, 8500, 12000, '', 1, 10, ''],
    ['网线 CAT6', 'NET-CAT6', '网络配件', '淘宝', '箱', '六类 305米', 5, 280, 450, '', 2, 20, ''],
]
write_sheet(ws1, '商品导入', product_headers, product_data,
            [18, 15, 12, 12, 8, 22, 8, 10, 10, 12, 10, 10, 20])

# ==========================================
#  Sheet 2: 供应商导入
# ==========================================
ws2 = wb.create_sheet()
supplier_headers = ['名称*', '联系人', '电话', '邮箱', '地址', '备注']
supplier_data = [
    ['华强五金', '张经理', '13800138001', 'zhang@huaqiang.com', '深圳市华强北路', ''],
    ['深华电子', '李工', '13900139002', '', '深圳市福田区', '电子元器件'],
    ['佳能', '', '', '', '', ''],
    ['淘宝', '', '', '', '', '线上采购'],
]
write_sheet(ws2, '供应商导入', supplier_headers, supplier_data,
            [20, 12, 16, 20, 30, 20])

# ==========================================
#  Sheet 3: 客户导入
# ==========================================
ws3 = wb.create_sheet()
customer_headers = ['名称*', '联系人', '电话', '邮箱', '地址', '备注']
customer_data = [
    ['中建公司', '王总', '13600136001', '', '北京市朝阳区', ''],
    ['华润集团', '', '', '', '', ''],
    ['个人客户A', '小明', '13700137001', '', '', '零售'],
]
write_sheet(ws3, '客户导入', customer_headers, customer_data,
            [20, 12, 16, 20, 30, 20])

# ==========================================
#  保存
# ==========================================
output_path = '库存数据导入模板.xlsx'
wb.save(output_path)
print(f"模板已生成: {output_path}")
print(f"  商品导入: {len(product_data)} 条示例")
print(f"  供应商导入: {len(supplier_data)} 条示例")
print(f"  客户导入: {len(customer_data)} 条示例")
print()
print("列名带 * 的必填，其他可留空。SKU 和库位不填会自动生成。")
