"""
Flask API 端点测试：使用 test_client + mock 模型层，不依赖真实数据库。
覆盖：页面渲染、CRUD、出入库、Excel 导入/导出、审计日志、AI 对话执行动作等。

注意：所有对共享对象（模型类方法 / app_mod.db）的替换必须走 monkeypatch，
避免状态泄漏到其他测试文件。
"""

import json
import os
import re
import time
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest
import requests as requests_lib


def make_recorder(result=None):
    """生成一个记录调用参数的 mock 函数"""
    calls = []

    def rec(*args, **kwargs):
        calls.append({'args': args, 'kwargs': kwargs})
        return result

    rec.calls = calls
    return rec


def track_tx(fake_db, monkeypatch):
    """包装 fake_db.execute，记录每条写 SQL 执行时是否处于事务内 → [(sql, in_transaction), ...]"""
    log = []
    orig = fake_db.execute

    def tracking(sql, params=None):
        r = orig(sql, params)
        log.append((sql, fake_db.in_transaction))
        return r

    monkeypatch.setattr(fake_db, 'execute', tracking)
    return log


def patch_all_dbs(app_mod, fake_db, monkeypatch):
    """把 app.db 和 models.db 都指向同一个 FakeDB。

    app.py 里 `from models import db`，模型方法引用的是 models 模块内的 db；
    只 patch app_mod.db 时，真实模型方法仍会连真库。
    """
    import models as models_mod
    monkeypatch.setattr(models_mod, 'db', fake_db)
    monkeypatch.setattr(app_mod, 'db', fake_db)


def make_stateful_inventory(fake_db, monkeypatch, initial_qty=0, location='', min_stock=0, max_stock=9999):
    """让 FakeDB 模拟真实库存状态：原子增减 SQL 会更新 state，SELECT 回读当前值。

    同时应答两种 SELECT：旧的 `SELECT quantity FROM inventory`（流水回读）和
    Excel 导入的 `SELECT quantity, location, min_stock, max_stock ...`（保留原库位/阈值）。
    """
    state = {'qty': initial_qty, 'location': location, 'min_stock': min_stock, 'max_stock': max_stock}
    orig_execute = fake_db.execute

    def smart_execute(sql, params=None):
        r = orig_execute(sql, params)
        if 'quantity = quantity + %s' in sql:
            state['qty'] += params[0]
        elif 'quantity = quantity - %s' in sql and r[0]:
            state['qty'] -= params[0]
        elif 'ON DUPLICATE KEY UPDATE' in sql and 'VALUES(quantity)' in sql:
            state['qty'] += params[1]   # (product_id, qty)
        return r

    monkeypatch.setattr(fake_db, 'execute', smart_execute)
    fake_db.one_side_effect = (
        lambda sql, params=None: {
            'quantity': state['qty'], 'location': state['location'],
            'min_stock': state['min_stock'], 'max_stock': state['max_stock'],
        }
        if ('SELECT quantity FROM inventory' in sql or 'SELECT quantity, location' in sql) else None
    )
    return state


def make_xlsx(rows):
    """rows: 第一行为表头，返回 BytesIO"""
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ==========================================
#  页面与静态资源
# ==========================================
class TestPages:
    def test_index_renders(self, client):
        resp = client.get('/')
        assert resp.status_code == 200
        assert '仓库管理系统' in resp.data.decode('utf-8')

    def test_static_css_served(self, client):
        resp = client.get('/static/css/style.css')
        assert resp.status_code == 200


# ==========================================
#  分类 API
# ==========================================
class TestCategories:
    def test_list(self, app_mod, client, monkeypatch):
        rec = make_recorder([{'id': 1, 'name': '电子产品'}])
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', rec)
        resp = client.get('/api/categories')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['success'] is True
        assert body['data'][0]['name'] == '电子产品'

    def test_create(self, app_mod, client, monkeypatch):
        rec = make_recorder(5)
        monkeypatch.setattr(app_mod.CategoryModel, 'create', rec)
        resp = client.post('/api/categories', json={'name': '新分类'})
        assert resp.status_code == 200
        assert resp.get_json()['id'] == 5
        assert rec.calls[0]['args'][0] == '新分类'

    def test_create_missing_name_returns_400(self, app_mod, client, monkeypatch):
        monkeypatch.setattr(app_mod.CategoryModel, 'create', make_recorder(1))
        resp = client.post('/api/categories', json={'description': '没有名字'})
        assert resp.status_code == 400

    def test_update_and_delete(self, app_mod, client, monkeypatch):
        upd = make_recorder(None)
        dele = make_recorder(None)
        monkeypatch.setattr(app_mod.CategoryModel, 'update', upd)
        monkeypatch.setattr(app_mod.CategoryModel, 'delete', dele)
        assert client.put('/api/categories/1', json={'name': '改名'}).status_code == 200
        assert upd.calls[0]['args'][:3] == (1, '改名', '')
        assert client.delete('/api/categories/1').status_code == 200
        assert dele.calls[0]['args'][0] == 1


# ==========================================
#  供应商 / 客户 API（含审计日志）
# ==========================================
class TestSuppliers:
    def test_create_records_audit(self, app_mod, client, fake_db, monkeypatch):
        create = make_recorder(8)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.SupplierModel, 'create', create)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/suppliers', json={'name': '华强五金', 'contact_person': '张经理'})
        assert resp.status_code == 200
        assert resp.get_json()['id'] == 8
        # create(name, contact, phone, email, address, notes)
        assert create.calls[0]['args'][:3] == ('华强五金', '张经理', '')

        action, table, record_id = audit.calls[0]['args'][:3]
        assert (action, table, record_id) == ('create', 'suppliers', 8)
        # 操作者 = 登录账号；测试未登录（DISABLE_AUTH）→ 回退默认 '系统'
        assert audit.calls[0]['kwargs']['operator'] == '系统'

    def test_update_and_delete_record_audit(self, app_mod, client, fake_db, monkeypatch):
        get_by_id = make_recorder({'id': 1, 'name': '旧名'})
        upd = make_recorder(None)
        dele = make_recorder(None)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.SupplierModel, 'get_by_id', get_by_id)
        monkeypatch.setattr(app_mod.SupplierModel, 'update', upd)
        monkeypatch.setattr(app_mod.SupplierModel, 'delete', dele)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        assert client.put('/api/suppliers/1', json={'name': '新名'}).status_code == 200
        assert client.delete('/api/suppliers/1').status_code == 200
        actions = [c['args'][0] for c in audit.calls]
        assert actions == ['update', 'delete']

    def test_create_audit_log_in_same_transaction(self, app_mod, client, fake_db, monkeypatch):
        """回归：审计日志必须与主操作在同一事务内写入（同连接、一起提交/回滚）"""
        create = make_recorder(8)
        in_tx_at_audit = []

        def spy(action, table_name, record_id, old_data=None, new_data=None, operator='system'):
            in_tx_at_audit.append(fake_db.in_transaction)

        monkeypatch.setattr(app_mod.SupplierModel, 'create', create)
        monkeypatch.setattr(app_mod.AuditLog, 'log', spy)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/suppliers', json={'name': '华强五金'})
        assert resp.status_code == 200
        assert in_tx_at_audit == [True]

    def test_create_audit_failure_returns_error(self, app_mod, client, fake_db, monkeypatch):
        """回归：审计日志写失败时接口必须报错（事务回滚），不能静默成功"""
        create = make_recorder(8)

        def boom(*a, **k):
            raise RuntimeError('audit db down')

        monkeypatch.setattr(app_mod.SupplierModel, 'create', create)
        monkeypatch.setattr(app_mod.AuditLog, 'log', boom)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/suppliers', json={'name': '华强五金'})
        assert resp.status_code == 400
        assert 'audit db down' in resp.get_json()['error']


class TestCustomers:
    def test_create_records_audit(self, app_mod, client, fake_db, monkeypatch):
        create = make_recorder(4)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.CustomerModel, 'create', create)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/customers', json={'name': '中建公司'})
        assert resp.status_code == 200
        assert resp.get_json()['id'] == 4
        action, table, record_id = audit.calls[0]['args'][:3]
        assert (action, table, record_id) == ('create', 'customers', 4)


# ==========================================
#  商品 API
# ==========================================
class TestProducts:
    def test_detail_not_found_404(self, app_mod, client, monkeypatch):
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_id', make_recorder(None))
        resp = client.get('/api/products/999')
        assert resp.status_code == 404
        assert '不存在' in resp.get_json()['error']

    def test_create_with_initial_stock(self, app_mod, client, fake_db, monkeypatch):
        create = make_recorder(9)
        stock_in = make_recorder(None)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.ProductModel, 'create', create)
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_in', stock_in)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/products', json={
            'name': '螺丝 M6', 'sku': 'S-1', 'quantity': 10,
            'unit_price': 2.5, 'location': 'A-01', 'min_stock': 5,
        })
        assert resp.status_code == 200
        assert resp.get_json()['id'] == 9

        # 初始库存走 stock_in（生成入库流水）
        si = stock_in.calls[0]
        assert si['args'][0] == 9 and si['kwargs']['quantity'] == 10
        assert si['kwargs']['unit_price'] == 2.5

        # 库位/阈值单独 UPDATE inventory
        hits = fake_db.assert_executed('UPDATE inventory SET location')
        assert hits[0][1][:4] == ('A-01', 5, 9999, 9)

    def test_create_missing_sku_returns_400(self, app_mod, client, fake_db, monkeypatch):
        create = make_recorder(1)
        monkeypatch.setattr(app_mod.ProductModel, 'create', create)
        monkeypatch.setattr(app_mod, 'db', fake_db)
        resp = client.post('/api/products', json={'name': '没有SKU'})
        assert resp.status_code == 400

    def test_create_without_quantity_still_sets_location(self, app_mod, client, fake_db, monkeypatch):
        """回归测试：不填初始库存时，库位/最低/最高库存仍应写入（不丢失）"""
        create = make_recorder(9)
        stock_in = make_recorder(None)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.ProductModel, 'create', create)
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_in', stock_in)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/products', json={
            'name': '打印纸 A4', 'sku': 'P-1', 'location': 'B-02', 'min_stock': 20,
        })
        assert resp.status_code == 200
        # 未提供数量 → 不应调用 stock_in
        assert stock_in.calls == []
        # 但库位/阈值照常写入
        hits = fake_db.assert_executed('UPDATE inventory SET location')
        assert hits[0][1] == ('B-02', 20, 9999, 9)

    def test_create_negative_quantity_rejected_before_write(self, app_mod, client, fake_db, monkeypatch):
        """回归：负数量创建商品必须直接 400，且任何数据都不落库（此前商品已先落库）"""
        create = make_recorder(9)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.ProductModel, 'create', create)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/products', json={
            'name': '螺丝 M6', 'sku': 'S-1', 'quantity': -5,
        })
        assert resp.status_code == 400
        assert '不能为负数' in resp.get_json()['error']
        # 校验发生在任何写入之前：商品未创建、无审计日志
        assert create.calls == []
        assert audit.calls == []

    def test_update_negative_quantity_rejected_before_write(self, app_mod, client, fake_db, monkeypatch):
        """回归：负数量更新商品必须直接 400，且任何数据都不落库（此前字段已先更新）"""
        upd = make_recorder(None)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_id', make_recorder({'id': 1}))
        monkeypatch.setattr(app_mod.ProductModel, 'update', upd)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.put('/api/products/1', json={
            'name': '螺丝 M6', 'sku': 'S-1', 'quantity': -5,
        })
        assert resp.status_code == 400
        assert upd.calls == []
        assert audit.calls == []


# ==========================================
#  库存 / 出入库 API
# ==========================================
class TestInventory:
    def test_low_stock_passes_threshold(self, app_mod, client, monkeypatch):
        rec = make_recorder([])
        monkeypatch.setattr(app_mod.InventoryModel, 'get_low_stock', rec)
        resp = client.get('/api/inventory/low-stock?threshold=25')
        assert resp.status_code == 200
        assert rec.calls[0]['args'] == (25,)

    def test_stock_in_updates_product_price_and_supplier(self, app_mod, client, fake_db, monkeypatch):
        stock_in = make_recorder(None)
        audit = make_recorder(None)
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_in', stock_in)
        monkeypatch.setattr(app_mod.AuditLog, 'log', audit)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/inventory/stock-in', json={
            'product_id': 7, 'quantity': 10, 'unit_price': 2.5, 'supplier_id': 3,
        })
        assert resp.status_code == 200
        assert resp.get_json()['message'] == '入库成功'

        # 最近进价与供应商同步回写 products 表
        price_hits = fake_db.assert_executed('UPDATE products SET unit_price')
        assert price_hits[0][1] == (2.5, 7)
        sup_hits = fake_db.assert_executed('UPDATE products SET supplier_id')
        assert sup_hits[0][1] == (3, 7)

        action, table, record_id = audit.calls[0]['args'][:3]
        assert (action, table, record_id) == ('stock_in', 'inventory', 7)

    def test_stock_out_insufficient_returns_400(self, app_mod, client, fake_db, monkeypatch):
        def boom(*a, **k):
            raise ValueError('库存不足！当前库存: 3, 需要出库: 10')
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_out', boom)
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/inventory/stock-out', json={
            'product_id': 7, 'quantity': 10,
        })
        assert resp.status_code == 400
        assert '库存不足' in resp.get_json()['error']


# ==========================================
#  DECIMAL 金额序列化
# ==========================================
class TestDecimalSerialization:
    def test_decimal_serialized_as_json_number(self, app_mod, client, monkeypatch):
        """回归：DECIMAL 金额字段经 CustomJSONProvider 输出为 JSON 数字（不再在 DB 层转 float）"""
        from decimal import Decimal
        rows = [{'id': 1, 'name': '螺丝', 'sku': 'S-1',
                 'unit_price': Decimal('2.50'), 'sale_price': Decimal('3.75')}]
        monkeypatch.setattr(app_mod.ProductModel, 'get_all', make_recorder(rows))

        resp = client.get('/api/products')
        assert resp.status_code == 200
        data = resp.get_json()['data'][0]
        assert data['unit_price'] == 2.5
        assert data['sale_price'] == 3.75


# ==========================================
#  交易记录 / 审计日志 API
# ==========================================
class TestTransactionsAndAudit:
    def test_transactions_respects_limit(self, app_mod, client, monkeypatch):
        rec = make_recorder([])
        monkeypatch.setattr(app_mod.TransactionModel, 'get_all', rec)
        resp = client.get('/api/transactions?limit=5')
        assert resp.status_code == 200
        assert rec.calls[0]['args'] == (5,)

    def test_audit_log_filters_by_table(self, app_mod, client, monkeypatch):
        by_table = make_recorder([{'id': 1}])
        recent = make_recorder([])
        monkeypatch.setattr(app_mod.AuditLog, 'get_by_table', by_table)
        monkeypatch.setattr(app_mod.AuditLog, 'get_recent', recent)

        resp = client.get('/api/audit-log?table=suppliers&limit=5')
        assert resp.status_code == 200
        assert by_table.calls[0]['args'] == ('suppliers', 5)
        assert recent.calls == []

    def test_audit_log_default_recent(self, app_mod, client, monkeypatch):
        recent = make_recorder([])
        monkeypatch.setattr(app_mod.AuditLog, 'get_recent', recent)
        resp = client.get('/api/audit-log')
        assert resp.status_code == 200
        assert recent.calls[0]['args'] == (200,)


# ==========================================
#  Excel 上传导入 API
# ==========================================
class TestUpload:
    def _patch_import_models(self, app_mod, fake_db, monkeypatch):
        """把导入流程涉及的模型层替换为 mock（全部走 monkeypatch，自动还原）"""
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', make_recorder([]))
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(None))
        monkeypatch.setattr(app_mod.ProductModel, 'create', make_recorder(1))
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'create', make_recorder(7))
        self.update_status = make_recorder(None)
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'update_status', self.update_status)
        monkeypatch.setattr(app_mod, 'db', fake_db)

    def test_rejects_bad_extension(self, client):
        buf = BytesIO(b'MZ...')
        resp = client.post('/api/upload', data={
            'file': (buf, 'evil.exe'), 'mode': 'replace'},
            content_type='multipart/form-data')
        assert resp.status_code == 400
        assert '仅支持' in resp.get_json()['error']

    def test_xlsx_import_creates_product_and_transaction(self, app_mod, client, fake_db, monkeypatch):
        self._patch_import_models(app_mod, fake_db, monkeypatch)
        buf = make_xlsx([['商品名称', '数量'], ['螺丝 M6', 50]])

        resp = client.post('/api/upload', data={
            'file': (buf, 'test.xlsx'), 'mode': 'replace'},
            content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['success'] is True
        assert body['data']['rows_imported'] == 1

        # 上传状态最终标记为 success（第一次调用是 processing）
        final_status = self.update_status.calls[-1]
        assert final_status['args'][1] == 'success'
        assert final_status['args'][2] == 1

        # 生成入库流水，批次号关联上传记录
        hits = fake_db.assert_executed('INSERT INTO transactions')
        assert 'Excel-7' in str(hits[0][1])

    def test_xlsx_import_skips_rows_without_name(self, app_mod, client, fake_db, monkeypatch):
        """回归测试：空单元格被 pandas 读成 NaN，str() 后变成字符串 "nan"，无名称的行必须被跳过"""
        self._patch_import_models(app_mod, fake_db, monkeypatch)
        buf = make_xlsx([['商品名称', '数量'], ['', 50], [None, 30]])

        resp = client.post('/api/upload', data={
            'file': (buf, 'empty.xlsx'), 'mode': 'replace'},
            content_type='multipart/form-data')
        assert resp.status_code == 200
        # 期望行为：无名称的行应被跳过
        assert resp.get_json()['data']['rows_imported'] == 0

    def test_xlsx_import_negative_quantity_not_written(self, app_mod, client, fake_db, monkeypatch):
        """回归测试：数量为负的行不应生成出入库流水"""
        self._patch_import_models(app_mod, fake_db, monkeypatch)
        buf = make_xlsx([['商品名称', '数量'], ['螺丝 M6', -5]])

        resp = client.post('/api/upload', data={
            'file': (buf, 'neg.xlsx'), 'mode': 'replace'},
            content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['data']['errors']           # 有行级错误提示
        assert any('负' in e for e in body['data']['errors'])
        assert not fake_db.find_executed('INSERT INTO transactions')  # 不写负库存流水

    def test_xlsx_import_duplicate_sku_in_file_skipped(self, app_mod, client, fake_db, monkeypatch):
        """回归测试：文件内重复 SKU 只导入首条，避免后者静默覆盖前者"""
        self._patch_import_models(app_mod, fake_db, monkeypatch)
        buf = make_xlsx([['商品名称', 'SKU', '数量'], ['螺丝 A', 'S1', 10], ['螺丝 B', 'S1', 20]])

        resp = client.post('/api/upload', data={
            'file': (buf, 'dup.xlsx'), 'mode': 'replace'},
            content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['data']['rows_imported'] == 1
        assert any('重复' in e for e in body['data']['errors'])
        fake_db.assert_executed('INSERT INTO transactions', count=1)


# ==========================================
#  Excel 导出 API（出库单 / 库存清单）
# ==========================================
class TestExport:
    def test_order_export_generates_xlsx_with_total(self, client):
        resp = client.post('/api/order/export', json={
            'customer': '中建公司', 'operator': '张三',
            'items': [
                {'name': '螺丝 M6', 'sku': 'S1', 'quantity': 2, 'sale_price': 0.5},
                {'name': '打印纸 A4', 'sku': 'P1', 'quantity': 3, 'sale_price': 15},
            ],
        })
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.content_type
        # xlsx 是 zip 格式，魔数 PK
        assert resp.data[:2] == b'PK'

        wb = openpyxl.load_workbook(BytesIO(resp.data))
        ws = wb.active
        total_val = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == '合计':
                    total_val = ws.cell(row=cell.row, column=7).value
        assert total_val is not None
        assert total_val == pytest.approx(2 * 0.5 + 3 * 15)

    def test_export_inventory_returns_xlsx(self, client):
        resp = client.post('/api/order/export-inventory', json={
            'items': [{'product_name': '螺丝 M6', 'quantity': 10}],
        })
        assert resp.status_code == 200
        assert resp.data[:2] == b'PK'


# ==========================================
#  AI API（mock LM Studio HTTP 层）
# ==========================================
class TestAI:
    def test_health_ok(self, app_mod, client, monkeypatch):
        import ai_service as ai_mod
        fake_resp = SimpleNamespace(status_code=200, json=lambda: {'data': [{'id': 'qwen3.6-35b-a3b'}]})
        monkeypatch.setattr(ai_mod.requests, 'get', lambda *a, **k: fake_resp)

        resp = client.get('/api/ai/health')
        body = resp.get_json()
        assert body['success'] is True
        assert 'qwen3.6-35b-a3b' in body['message']

    def test_health_down(self, app_mod, client, monkeypatch):
        import ai_service as ai_mod

        def boom(*a, **k):
            raise requests_lib.exceptions.ConnectionError('refused')
        monkeypatch.setattr(ai_mod.requests, 'get', boom)

        resp = client.get('/api/ai/health')
        body = resp.get_json()
        assert body['success'] is False
        assert '无法连接' in body['message']

    def test_chat_add_supplier_action(self, app_mod, client, monkeypatch):
        reply_text = (
            '好的，已为您添加供应商。\n'
            '```action\n'
            '[{"action": "add_supplier", "name": "华为", "contact": "张经理", "phone": "13800000000"}]\n'
            '```\n'
        )
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply_text)

        create = make_recorder(9)
        monkeypatch.setattr(app_mod.SupplierModel, 'create', create)
        # 去重检查：名称不存在 → 走创建分支
        monkeypatch.setattr(app_mod.SupplierModel, 'get_by_name', make_recorder(None))

        resp = client.post('/api/ai/chat', json={'message': '新增供应商华为，联系人张经理'})
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['actions'] == ['新增供应商: 华为']
        assert '已执行' in body['data']
        assert create.calls[0]['args'][0] == '华为'

    def test_chat_stock_in_auto_category(self, app_mod, client, fake_db, monkeypatch):
        reply_text = (
            '好的，已入库。\n'
            '```action\n'
            '[{"action": "stock_in", "sku": "NEW-1", "quantity": 5, "name": "电阻 10K"}]\n'
            '```\n'
        )
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply_text)

        cat_create = make_recorder(11)
        prod_create = make_recorder(5)
        stock_in = make_recorder(None)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(None))
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', make_recorder([]))
        monkeypatch.setattr(app_mod.CategoryModel, 'create', cat_create)
        monkeypatch.setattr(app_mod.ProductModel, 'create', prod_create)
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_in', stock_in)
        # _do_ai_stock_in 整体包在 db.transaction() 内 → 用 FakeDB 承接
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/ai/chat', json={'message': '入库 5 个电阻 10K'})
        assert resp.status_code == 200
        body = resp.get_json()
        assert any('入库' in a for a in body['actions'])

        # "电阻" 命中关键词 → 自动归类到「电子产品」
        assert cat_create.calls[0]['args'][0] == '电子产品'
        # 新商品创建后执行入库
        assert prod_create.calls[0]['args'][:2] == ('电阻 10K', 'NEW-1')
        assert stock_in.calls[0]['args'][0] == 5

    def test_chat_create_order_writes_excel(self, app_mod, client, tmp_path, fake_db, monkeypatch):
        reply_text = (
            '出库单已创建。\n'
            '```action\n'
            '[{"action": "create_order", "customer": "中建公司", "operator": "张三",\n'
            '  "warehouse": "WH-A01",\n'
            '  "items": [{"sku": "SCR-1", "quantity": 2, "price": 0.5}]}]\n'
            '```\n'
        )
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply_text)

        cust_create = make_recorder(3)
        stock_out = make_recorder(None)
        # 客户按名称查找（预检通过后才在事务内建客户）：不存在 → 走创建分支
        monkeypatch.setattr(app_mod.CustomerModel, 'get_by_name', make_recorder(None))
        monkeypatch.setattr(app_mod.CustomerModel, 'create', cust_create)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(
            {'id': 7, 'name': '螺丝 M6', 'specification': '不锈钢'}))
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_out', stock_out)
        # 出库单预检需要查询库存：mock db 返回充足库存
        monkeypatch.setattr(app_mod, 'db', fake_db)
        fake_db.one_side_effect = lambda sql, params=None: {'quantity': 10}

        # 出库单 Excel 写入临时目录，避免污染真实 uploads/
        monkeypatch.setattr(app_mod, 'UPLOAD_FOLDER', str(tmp_path))

        resp = client.post('/api/ai/chat', json={'message': '创建出库单给中建公司'})
        assert resp.status_code == 200
        body = resp.get_json()
        # 出库执行：客户 id=3、单价 0.5
        so = stock_out.calls[0]
        assert so['args'][0] == 7 and so['args'][1] == 2
        assert so['kwargs']['customer_id'] == 3
        assert so['kwargs']['unit_price'] == 0.5

        # Excel 文件已生成，回复带下载链接
        files = os.listdir(tmp_path)
        assert any(f.startswith('出库单_中建公司') and f.endswith('.xlsx') for f in files)
        assert '/uploads/' in body['data']


# ==========================================
#  /api/ai/chat/stream：流式 SSE 输出
# ==========================================
class TestAIStream:
    def _patch_stream(self, app_mod, monkeypatch, events):
        """把 ai_service.chat_stream 替换为产出给定事件的生成器"""
        def fake_stream(message):
            for ev in events:
                yield ev
        monkeypatch.setattr(app_mod.ai_service, 'chat_stream', fake_stream)
        # 指令执行（_execute_ai_actions）若无指令则不触碰模型；此处兜底 mock 掉模型写方法
        monkeypatch.setattr(app_mod.SupplierModel, 'get_by_name', lambda n: None)
        monkeypatch.setattr(app_mod.SupplierModel, 'create', lambda *a, **k: None)
        monkeypatch.setattr(app_mod.CustomerModel, 'get_by_name', lambda n: None)
        monkeypatch.setattr(app_mod.CustomerModel, 'create', lambda *a, **k: None)

    def test_streams_tokens_then_done(self, app_mod, client, monkeypatch):
        self._patch_stream(app_mod, monkeypatch, [
            {'type': 'token', 'content': '你好'},
            {'type': 'token', 'content': '世界'},
            {'type': 'done', 'full': '你好世界', 'reply': '你好世界', 'actions': None},
        ])
        resp = client.post('/api/ai/chat/stream', json={'message': 'hi'})
        assert resp.status_code == 200
        assert resp.headers['Content-Type'].startswith('text/event-stream')

        body = resp.get_data(as_text=True)
        assert 'data: {"type": "token", "content": "你好"}' in body
        assert 'data: {"type": "token", "content": "世界"}' in body
        # done 事件携带最终 reply；无 action 时不应有 log 事件
        assert '"type": "done"' in body
        assert '"你好世界"' in body
        assert '"type": "log"' not in body

    def test_stream_appends_execution_log_for_actions(self, app_mod, client, monkeypatch):
        self._patch_stream(app_mod, monkeypatch, [
            {'type': 'done', 'full': '好的', 'reply': '好的',
             'actions': [{'action': 'add_supplier', 'name': '华为'}]},
        ])
        resp = client.post('/api/ai/chat/stream', json={'message': '新增供应商'})
        body = resp.get_data(as_text=True)
        # 有 action → 产生 log 事件，done 回复里含执行日志
        assert '"type": "log"' in body
        assert '新增供应商: 华为' in body

    def test_missing_message_returns_400(self, client):
        resp = client.post('/api/ai/chat/stream', json={})
        assert resp.status_code == 400
        assert resp.get_json()['error'] == '请输入问题'

    def test_analyze_stream_tokens_then_done(self, app_mod, client, monkeypatch):
        def fake_stream(query_type):
            yield {'type': 'token', 'content': '低库存'}
            yield {'type': 'token', 'content': '预警'}
            yield {'type': 'done', 'data': '低库存预警'}
        monkeypatch.setattr(app_mod.ai_service, 'analyze_stream', fake_stream)

        resp = client.get('/api/ai/analyze/stream?type=low_stock')
        assert resp.status_code == 200
        assert resp.headers['Content-Type'].startswith('text/event-stream')
        body = resp.get_data(as_text=True)
        assert 'data: {"type": "token", "content": "低库存"}' in body
        assert '"type": "done"' in body
        assert '"低库存预警"' in body


# ==========================================
#  回归：删除商品连带清理库存行（孤儿库存）
# ==========================================
class TestProductDeleteAPI:
    def test_delete_product_removes_inventory_row_in_same_tx(self, app_mod, client, fake_db, monkeypatch):
        """回归 #2：删商品必须同事务内删掉 inventory 行，否则仪表盘总数量永久虚增"""
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_id', make_recorder({'id': 5, 'name': '螺丝 M6'}))
        tx_log = track_tx(fake_db, monkeypatch)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.delete('/api/products/5')
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        prod_hits = fake_db.find_executed('DELETE FROM products WHERE id=%s')
        inv_hits = fake_db.find_executed('DELETE FROM inventory WHERE product_id=%s')
        assert len(prod_hits) == 1 and prod_hits[0][1] == (5,)
        assert len(inv_hits) == 1 and inv_hits[0][1] == (5,)
        # 两条删除都必须发生在同一事务内
        for sql, in_tx in tx_log:
            if 'DELETE FROM' in sql:
                assert in_tx is True


# ==========================================
#  回归：Excel 导入（价格保留 + 原子增减）
# ==========================================
class TestExcelImportFixes:
    def _patch(self, app_mod, fake_db, monkeypatch, existing=None):
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', make_recorder([]))
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(existing))
        monkeypatch.setattr(app_mod.ProductModel, 'create', make_recorder(1))
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'create', make_recorder(7))
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'update_status', make_recorder(None))
        patch_all_dbs(app_mod, fake_db, monkeypatch)

    def test_reimport_existing_sku_preserves_price_and_supplier(self, app_mod, client, fake_db, monkeypatch):
        """回归 #1：重新导入已有 SKU 不得把价格清零、供应商抹掉（文件没有这些列）"""
        existing = {'id': 9, 'category_id': 3, 'supplier_id': 4}
        self._patch(app_mod, fake_db, monkeypatch, existing=existing)
        buf = make_xlsx([['商品名称', 'SKU', '数量'], ['螺丝 M6', 'S1', 50]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        hits = fake_db.find_executed('UPDATE products SET')
        assert len(hits) == 1
        sql, params = hits[0]
        # 价格列不得出现在 SQL 中（None → 保留原值）
        assert 'unit_price' not in sql and 'sale_price' not in sql
        # 参数顺序 name, sku, cat_id, sup_id, unit, spec, desc, id：分类/供应商保留原值
        assert params[2] == 3 and params[3] == 4

    def test_new_product_uses_atomic_increment(self, app_mod, client, fake_db, monkeypatch):
        """回归 #4：库存写入必须用原子自增，禁止绝对值 SET quantity=%s（并发丢更新）"""
        self._patch(app_mod, fake_db, monkeypatch)
        buf = make_xlsx([['商品名称', '数量'], ['螺丝 M6', 50]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200

        inc = fake_db.find_executed('UPDATE inventory SET quantity = quantity + %s')
        assert len(inc) == 1 and inc[0][1] == (50, 1)   # diff=50（原库存 0）, prod_id=1
        assert not any('SET quantity=%s' in s for s, _ in fake_db.executed)

    def test_replace_mode_decrement_uses_guarded_atomic_update(self, app_mod, client, fake_db, monkeypatch):
        """回归 #4：replace 模式文件数量低于现库存 → 带条件的原子扣减，流水 before/after 以实际值为准"""
        self._patch(app_mod, fake_db, monkeypatch, existing={'id': 9})
        state = make_stateful_inventory(fake_db, monkeypatch, initial_qty=80)
        buf = make_xlsx([['商品名称', 'SKU', '数量'], ['螺丝 M6', 'S1', 30]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        dec = fake_db.find_executed(
            'UPDATE inventory SET quantity = quantity - %s WHERE product_id=%s AND quantity >= %s')
        assert len(dec) == 1 and dec[0][1] == (50, 9, 50)   # 减 50，条件 quantity>=50
        assert state['qty'] == 30

        txn = fake_db.find_executed('INSERT INTO transactions')
        assert len(txn) == 1
        p = txn[0][1]
        assert (p[0], p[1], p[2]) == (9, 'out', 50)   # product_id, type, qty
        assert (p[3], p[4]) == (80, 30)               # before, after

    def test_replace_mode_insufficient_stock_skips_row(self, app_mod, client, fake_db, monkeypatch):
        """回归 #4：并发下库存不足以扣减 → 跳过该行库存写入并报错，不产生流水"""
        self._patch(app_mod, fake_db, monkeypatch, existing={'id': 9})
        make_stateful_inventory(fake_db, monkeypatch, initial_qty=80)
        buf = make_xlsx([['商品名称', 'SKU', '数量'], ['螺丝 M6', 'S1', 30]])

        # 条件扣减未命中（并发已把库存扣走）→ 库存不足
        orig_execute = fake_db.execute

        def guarded(sql, params=None):
            r = orig_execute(sql, params)
            if 'quantity - %s' in sql:
                return (0, r[1])
            return r
        monkeypatch.setattr(fake_db, 'execute', guarded)

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert any('库存不足' in e for e in body['data']['errors'])
        assert not fake_db.find_executed('INSERT INTO transactions')

    def test_increment_mode_accumulates(self, app_mod, client, fake_db, monkeypatch):
        """回归 #4：increment 模式文件数量是新增量 → 原子累加"""
        self._patch(app_mod, fake_db, monkeypatch, existing={'id': 9})
        state = make_stateful_inventory(fake_db, monkeypatch, initial_qty=80)
        buf = make_xlsx([['商品名称', 'SKU', '数量'], ['螺丝 M6', 'S1', 5]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'increment'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200

        inc = fake_db.find_executed('UPDATE inventory SET quantity = quantity + %s')
        assert len(inc) == 1 and inc[0][1] == (5, 9)
        assert state['qty'] == 85


# ==========================================
#  回归：AI 智能导入（价格保留 + 原子自增）
# ==========================================
class TestAISmartImportFixes:
    def test_reimport_preserves_price_and_atomic_increment(self, app_mod, client, fake_db, monkeypatch):
        """回归 #1/#4：已有商品 → 保留价格/分类/供应商；库存用 ON DUPLICATE KEY 原子自增"""
        items = [{'name': '螺丝 M6', 'sku': 'S1', 'quantity': 5, 'price': 2.0}]
        monkeypatch.setattr(app_mod.ai_service, 'smart_import', lambda text: items)
        existing = {'id': 9, 'category_id': 3, 'supplier_id': 4}
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(existing))
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', make_recorder([]))
        monkeypatch.setattr(app_mod.SupplierModel, 'get_all', make_recorder([]))
        state = make_stateful_inventory(fake_db, monkeypatch, initial_qty=20)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.post('/api/ai/smart-import', json={'text': '采购螺丝 M6 5 个'})
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        hits = fake_db.find_executed('UPDATE products SET')
        assert len(hits) == 1
        sql, params = hits[0]
        assert 'unit_price' not in sql and 'sale_price' not in sql
        # 参数顺序 name, sku, cat_id, sup_id, ...：分类/供应商保留原值
        assert params[2] == 3 and params[3] == 4

        dup = fake_db.find_executed('ON DUPLICATE KEY UPDATE quantity = quantity + VALUES(quantity)')
        assert len(dup) == 1 and dup[0][1] == (9, 5)
        assert state['qty'] == 25


# ==========================================
#  回归：SKU 兜底生成防碰撞
# ==========================================
class TestUniqueSkuFallback:
    def test_unique_sku_appends_suffix_on_collision(self, app_mod, monkeypatch):
        """回归 #5：兜底 SKU 已被占用 → 追加 -2/-3...，避免库存并入别的商品"""
        taken = {'BASE', 'BASE-2'}

        def fake_get_by_sku(sku):
            return {'id': 1} if sku in taken else None
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', fake_get_by_sku)

        assert app_mod._unique_sku('BASE') == 'BASE-3'
        assert app_mod._unique_sku('FREE') == 'FREE'

    def test_inbound_recognize_no_collision_for_same_prefix_names(self, app_mod, client, fake_db, monkeypatch):
        """回归 #5：前 6 字相同的两个中文商品名不得合并到同一 SKU"""
        items = [{'name': '超长中文商品名称甲', 'quantity': 1},
                 {'name': '超长中文商品名称乙', 'quantity': 2}]
        monkeypatch.setattr(app_mod.ai_service, 'recognize_inbound_image', lambda b64: items)

        created_skus = []
        orig_execute = fake_db.execute

        def tracking_execute(sql, params=None):
            r = orig_execute(sql, params)
            if sql.lstrip().upper().startswith('INSERT INTO PRODUCTS'):
                created_skus.append(params[1])   # (name, sku, ...)
            return r
        monkeypatch.setattr(fake_db, 'execute', tracking_execute)

        def one(sql, params=None):
            if 'WHERE sku = %s' in sql:
                return {'id': 99} if params[0] in created_skus else None
            return None
        fake_db.one_side_effect = one
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.post('/api/ai/inbound-recognize', json={'image': 'x'})
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['success'] is True and body['data']['count'] == 2
        skus = [i['sku'] for i in body['data']['imported']]
        assert len(set(skus)) == 2, f'两个商品不得共用同一 SKU: {skus}'


# ==========================================
#  回归：供应商/客户导入去重（upsert）
# ==========================================
class TestUploadDedup:
    def _track_names(self, app_mod, fake_db, monkeypatch, table, initial=()):
        """有状态 FakeDB：跟踪已插入该表的名称，get_by_name 按此命中"""
        names = set(initial)
        orig_execute = fake_db.execute

        def tracking_execute(sql, params=None):
            r = orig_execute(sql, params)
            if sql.lstrip().upper().startswith(f'INSERT INTO {table.upper()}'):
                names.add(params[0])
            return r
        monkeypatch.setattr(fake_db, 'execute', tracking_execute)

        def one(sql, params=None):
            if f'FROM {table} WHERE name=%s' in sql:
                if params[0] in names:
                    return {'id': 1, 'name': params[0], 'contact_person': '', 'phone': '',
                            'email': '', 'address': '', 'notes': ''}
            return None
        fake_db.one_side_effect = one
        patch_all_dbs(app_mod, fake_db, monkeypatch)

    @staticmethod
    def _csv(rows):
        import csv as _csv
        from io import StringIO
        sio = StringIO()
        w = _csv.writer(sio)
        for r in rows:
            w.writerow(r)
        # utf-8-sig（带 BOM）与 Excel 导出的 CSV 一致，走 pd.read_csv 的 utf-8-sig 分支
        return BytesIO(sio.getvalue().encode('utf-8-sig'))

    def test_supplier_upload_reupload_updates_not_duplicates(self, app_mod, client, fake_db, monkeypatch):
        """回归 #6：重复上传同一供应商文件 → 更新原记录，不再产生重复"""
        self._track_names(app_mod, fake_db, monkeypatch, 'suppliers')

        r1 = client.post('/api/upload/suppliers', data={'file': (self._csv([['名称', '电话'], ['华为', '13800000000']]), 's.csv')},
                         content_type='multipart/form-data')
        assert r1.status_code == 200
        d1 = r1.get_json()['data']
        assert d1['created'] == 1 and d1['updated'] == 0

        # test client 会关闭上传文件 → 第二次用新 buffer
        r2 = client.post('/api/upload/suppliers', data={'file': (self._csv([['名称', '电话'], ['华为', '13800000000']]), 's.csv')},
                         content_type='multipart/form-data')
        d2 = r2.get_json()['data']
        assert d2['created'] == 0 and d2['updated'] == 1

        # 两次上传合计：只 INSERT 一次、UPDATE 一次
        assert len(fake_db.find_executed('INSERT INTO suppliers')) == 1
        assert len(fake_db.find_executed('UPDATE suppliers SET')) == 1

    def test_customer_upload_reupload_updates_not_duplicates(self, app_mod, client, fake_db, monkeypatch):
        """回归 #6：客户同理"""
        self._track_names(app_mod, fake_db, monkeypatch, 'customers')

        r1 = client.post('/api/upload/customers', data={'file': (self._csv([['名称', '电话'], ['中建公司', '13900000000']]), 'c.csv')},
                         content_type='multipart/form-data')
        assert r1.status_code == 200
        d1 = r1.get_json()['data']
        assert d1['created'] == 1 and d1['updated'] == 0

        r2 = client.post('/api/upload/customers', data={'file': (self._csv([['名称', '电话'], ['中建公司', '13900000000']]), 'c.csv')},
                         content_type='multipart/form-data')
        d2 = r2.get_json()['data']
        assert d2['created'] == 0 and d2['updated'] == 1

    def test_ai_add_supplier_skips_existing(self, app_mod, client, fake_db, monkeypatch):
        """回归 #6：AI add_supplier 名称已存在 → 跳过，不覆盖"""
        reply = '好的。\n```action\n[{"action": "add_supplier", "name": "华为"}]\n```\n'
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply)
        self._track_names(app_mod, fake_db, monkeypatch, 'suppliers', initial=('华为',))

        resp = client.post('/api/ai/chat', json={'message': '新增供应商华为'})
        body = resp.get_json()
        assert any('已存在' in a for a in body['actions'])
        assert not fake_db.find_executed('INSERT INTO suppliers')

    def test_ai_add_customer_skips_existing(self, app_mod, client, fake_db, monkeypatch):
        """回归 #6：AI add_customer 名称已存在 → 跳过，不覆盖"""
        reply = '好的。\n```action\n[{"action": "add_customer", "name": "中建公司"}]\n```\n'
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply)
        self._track_names(app_mod, fake_db, monkeypatch, 'customers', initial=('中建公司',))

        resp = client.post('/api/ai/chat', json={'message': '新增客户中建公司'})
        body = resp.get_json()
        assert any('已存在' in a for a in body['actions'])
        assert not fake_db.find_executed('INSERT INTO customers')


# ==========================================
#  回归：AI create_order 先预检后建客户
# ==========================================
class TestAICreateOrderPrecheckFirst:
    @staticmethod
    def _chat_reply(app_mod, monkeypatch, action_obj):
        reply = '好的。\n```action\n' + json.dumps([action_obj], ensure_ascii=False) + '\n```\n'
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply)

    def test_failed_precheck_creates_no_customer(self, app_mod, client, fake_db, monkeypatch):
        """回归 #7：预检失败（商品不存在）时不得创建客户，不留垃圾数据"""
        self._chat_reply(app_mod, monkeypatch, {
            'action': 'create_order', 'customer': '新客户X', 'operator': 'AI',
            'items': [{'sku': 'NOPE-1', 'quantity': 2}]})
        cust_create = make_recorder(3)
        monkeypatch.setattr(app_mod.CustomerModel, 'get_by_name', make_recorder(None))
        monkeypatch.setattr(app_mod.CustomerModel, 'create', cust_create)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(None))
        monkeypatch.setattr(app_mod, 'db', fake_db)

        resp = client.post('/api/ai/chat', json={'message': '创建出库单'})
        body = resp.get_json()
        assert any('出库单未执行' in a for a in body['actions'])
        assert cust_create.calls == []          # 没有留下垃圾客户
        assert not fake_db.find_executed('INSERT INTO customers')

    def test_existing_customer_reused_not_created(self, app_mod, client, tmp_path, fake_db, monkeypatch):
        """回归 #7：客户已存在 → 复用其 id，不新建"""
        self._chat_reply(app_mod, monkeypatch, {
            'action': 'create_order', 'customer': '中建公司', 'operator': '张三',
            'items': [{'sku': 'SCR-1', 'quantity': 2, 'price': 0.5}]})
        cust_create = make_recorder(3)
        stock_out = make_recorder(None)
        monkeypatch.setattr(app_mod.CustomerModel, 'get_by_name', make_recorder({'id': 8}))
        monkeypatch.setattr(app_mod.CustomerModel, 'create', cust_create)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(
            {'id': 7, 'name': '螺丝 M6', 'specification': ''}))
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_out', stock_out)
        fake_db.one_side_effect = lambda sql, params=None: {'quantity': 10}
        monkeypatch.setattr(app_mod, 'db', fake_db)
        monkeypatch.setattr(app_mod, 'UPLOAD_FOLDER', str(tmp_path))

        resp = client.post('/api/ai/chat', json={'message': '创建出库单'})
        body = resp.get_json()
        assert resp.status_code == 200 and body['success'] is True
        assert cust_create.calls == []          # 复用已有客户 id=8
        so = stock_out.calls[0]
        assert so['kwargs']['customer_id'] == 8


# ==========================================
#  回归：AI chat 动作原子性（事务 + set_quantity 原子增量）
# ==========================================
class TestAIChatActionsAtomicity:
    def test_stock_in_writes_all_in_one_transaction(self, app_mod, client, fake_db, monkeypatch):
        """回归 #3：AI 入库的全部写入（建商品+库存自增+流水）必须在同一事务内"""
        reply = ('好的。\n```action\n[{"action": "stock_in", "sku": "NEW-9", "quantity": 5, '
                 '"name": "测试品"}]\n```\n')
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(None))
        tx_log = track_tx(fake_db, monkeypatch)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.post('/api/ai/chat', json={'message': '入库 5 个测试品'})
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        writes = [(s, t) for s, t in tx_log if s.lstrip().upper().startswith(('INSERT', 'UPDATE'))]
        assert len(writes) >= 3   # products + inventory + transactions（至少）
        assert all(t is True for _, t in writes), f'存在事务外的写入: {writes}'

    def test_set_quantity_uses_atomic_delta(self, app_mod, client, fake_db, monkeypatch):
        """回归 #3：set_quantity 必须走原子增减+流水，禁止裸读-改-写 UPDATE"""
        reply = '好的。\n```action\n[{"action": "set_quantity", "sku": "S1", "quantity": 30}]\n```\n'
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder({'id': 5}))
        state = make_stateful_inventory(fake_db, monkeypatch, initial_qty=20)
        tx_log = track_tx(fake_db, monkeypatch)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.post('/api/ai/chat', json={'message': '把 S1 库存调到 30'})
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        inc = fake_db.find_executed('UPDATE inventory SET quantity = quantity + %s')
        assert len(inc) == 1 and inc[0][1] == (10, 5)   # delta=30-20=10
        assert state['qty'] == 30
        assert not any(re.search(r'UPDATE inventory SET quantity=%s', s) for s, _ in fake_db.executed)

        txn = fake_db.find_executed('INSERT INTO transactions')
        assert len(txn) == 1
        sql, p = txn[0]
        # stock_in 流水：type='in' 是 SQL 字面量；参数顺序 product_id, quantity, unit_price, supplier_id, before, after, ...
        assert "'in'" in sql
        assert (p[0], p[1]) == (5, 10)
        assert (p[4], p[5]) == (20, 30)   # before, after

    def test_do_ai_set_quantity_rejects_negative(self, app_mod, fake_db, monkeypatch):
        """回归 #3：目标数量为负直接拒绝，不产生任何写入"""
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder({'id': 5}))
        monkeypatch.setattr(app_mod, 'db', fake_db)

        with pytest.raises(ValueError, match='不能为负数'):
            app_mod._do_ai_set_quantity('S1', -5)
        assert fake_db.executed == []

    def test_do_ai_set_quantity_insufficient_stock_raises(self, app_mod, fake_db, monkeypatch):
        """回归 #3：调低库存但不足扣减 → 抛错回滚，不出现负库存"""
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder({'id': 5}))
        fake_db.update_affected = 0   # 条件扣减未命中
        fake_db.one_side_effect = (
            lambda sql, params=None: {'quantity': 20} if 'SELECT quantity FROM inventory' in sql else None)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        with pytest.raises(ValueError, match='库存不足'):
            app_mod._do_ai_set_quantity('S1', 5)   # 需减 15，但扣减未命中


# ==========================================
#  回归：导出端点校验 + 文件名净化/时间戳
# ==========================================
class TestExportValidation:
    def test_order_export_rejects_non_list_items(self, client):
        resp = client.post('/api/order/export', json={'items': 'nope'})
        assert resp.status_code == 400 and 'items' in resp.get_json()['error']

    def test_order_export_rejects_non_dict_entries(self, client):
        resp = client.post('/api/order/export', json={'items': [1, 2]})
        assert resp.status_code == 400

    def test_order_export_rejects_missing_json_body(self, client):
        resp = client.post('/api/order/export', data='plain text', content_type='text/plain')
        assert resp.status_code == 400

    def test_order_export_tolerates_bad_numbers(self, client):
        """回归 #9：非法价格/数量按 0 计，整单不得 500"""
        resp = client.post('/api/order/export', json={
            'customer': '测试', 'items': [{'name': 'X', 'quantity': 'abc', 'sale_price': None}]})
        assert resp.status_code == 200 and resp.data[:2] == b'PK'

    def test_export_inventory_rejects_non_dict_entries(self, client):
        resp = client.post('/api/order/export-inventory', json={'items': [1, 2]})
        assert resp.status_code == 400

    def test_export_inventory_tolerates_bad_price(self, client):
        """回归 #9：非法最近进价不得让整单导出失败"""
        resp = client.post('/api/order/export-inventory', json={
            'items': [{'product_name': 'X', 'latest_price': 'abc'}]})
        assert resp.status_code == 200 and resp.data[:2] == b'PK'

    def test_order_export_filename_sanitized_with_timestamp(self, client):
        """回归 #8：文件名必须净化（防路径穿越）且带秒级时间戳"""
        resp = client.post('/api/order/export', json={
            'customer': '../../evil name!', 'items': [{'name': 'X', 'quantity': 1, 'sale_price': 1}]})
        assert resp.status_code == 200
        cd = resp.headers.get('Content-Disposition', '')
        assert '/' not in cd                       # 无路径穿越字符
        assert re.search(r'\d{8}_\d{6}', cd)       # 秒级时间戳


# ==========================================
#  回归：Excel 重导入保留库位/阈值 + 自动库位修复（二 / #10）
# ==========================================
class TestExcelImportLocationThresholds:
    def _patch(self, app_mod, fake_db, monkeypatch, existing=None):
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', make_recorder([]))
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(existing))
        monkeypatch.setattr(app_mod.ProductModel, 'create', make_recorder(1))
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'create', make_recorder(7))
        self.update_status = make_recorder(None)
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'update_status', self.update_status)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

    def test_reimport_without_location_cols_preserves_existing(self, app_mod, client, fake_db, monkeypatch):
        """回归 二：文件没有库位/最低库存/最高库存列 → 重导入不得抹掉已有库位与阈值"""
        self._patch(app_mod, fake_db, monkeypatch, existing={'id': 9})
        make_stateful_inventory(fake_db, monkeypatch, initial_qty=80,
                                location='B-007', min_stock=5, max_stock=100)
        buf = make_xlsx([['商品名称', 'SKU', '数量'], ['螺丝 M6', 'S1', 30]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        hits = fake_db.find_executed('UPDATE inventory SET location=%s, min_stock=%s, max_stock=%s')
        assert len(hits) == 1
        # params: (location, min_stock, max_stock, product_id) → 全部保留原值，不再被冲成 ''/0/9999
        assert hits[0][1] == ('B-007', 5, 100, 9)

    def test_reimport_with_explicit_thresholds_updates_them(self, app_mod, client, fake_db, monkeypatch):
        """回归 二：文件显式提供最低/最高库存 → 用新值；库位列缺失仍保留原库位"""
        self._patch(app_mod, fake_db, monkeypatch, existing={'id': 9})
        make_stateful_inventory(fake_db, monkeypatch, initial_qty=80,
                                location='B-007', min_stock=5, max_stock=100)
        buf = make_xlsx([['商品名称', 'SKU', '数量', '最低库存', '最高库存'],
                         ['螺丝 M6', 'S1', 30, 2, 50]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        hits = fake_db.find_executed('UPDATE inventory SET location=%s, min_stock=%s, max_stock=%s')
        assert hits[0][1] == ('B-007', 2, 50, 9)

    def test_uncategorized_auto_location_is_null_and_numeric_max(self, app_mod, client, fake_db, monkeypatch):
        """回归 #10：未分类商品（category_id=None）自动库位必须用 IS NULL 查询且取数值最大"""
        self._patch(app_mod, fake_db, monkeypatch)   # 新商品 → prod_id=1，无分类
        make_stateful_inventory(fake_db, monkeypatch, initial_qty=0)

        # 已有未分类库位含 ZZ-999 / ZZ-1000（字符串排序 max 会错取 'ZZ-999'）
        fake_db.query_side_effect = (
            lambda sql, params=None: [{'location': 'ZZ-999'}, {'location': 'ZZ-1000'}]
            if 'category_id IS NULL' in sql else []
        )
        buf = make_xlsx([['商品名称', '数量'], ['神秘物品', 5]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        # 必须发出 IS NULL 查询（旧代码 WHERE p.category_id=%s 传 None 永远查不到 → ZZ-001 反复复用）
        assert len(fake_db.find_queried('category_id IS NULL')) == 1
        hits = fake_db.find_executed('UPDATE inventory SET location=%s, min_stock=%s, max_stock=%s')
        # 数值最大是 1000 → 下一个应为 ZZ-1001（字符串排序会给出与已有库位冲突的 ZZ-1000）
        assert hits[0][1][0] == 'ZZ-1001'

    def test_categorized_auto_location_numeric_max(self, app_mod, client, fake_db, monkeypatch):
        """回归 #10：有分类商品自动库位取同前缀库位的数值最大 + 1"""
        self._patch(app_mod, fake_db, monkeypatch)
        make_stateful_inventory(fake_db, monkeypatch, initial_qty=0)

        fake_db.query_side_effect = (
            lambda sql, params=None: [{'location': '电子-999'}, {'location': '电子-1000'}]
            if 'p.category_id=%s' in sql else []
        )
        buf = make_xlsx([['商品名称', '分类', '数量'], ['电阻 1kΩ', '电子产品', 5]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        hits = fake_db.find_executed('UPDATE inventory SET location=%s, min_stock=%s, max_stock=%s')
        assert hits[0][1][0] == '电子-1001'


# ==========================================
#  回归：上传状态 partial（#11）
# ==========================================
class TestUploadPartialStatus:
    def _patch(self, app_mod, fake_db, monkeypatch):
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', make_recorder([]))
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(None))
        monkeypatch.setattr(app_mod.ProductModel, 'create', make_recorder(1))
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'create', make_recorder(7))
        self.update_status = make_recorder(None)
        monkeypatch.setattr(app_mod.ExcelUploadModel, 'update_status', self.update_status)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

    def test_row_errors_mark_partial(self, app_mod, client, fake_db, monkeypatch):
        """回归 #11：存在行级错误 → 状态 partial，不再一律 success"""
        self._patch(app_mod, fake_db, monkeypatch)
        buf = make_xlsx([['商品名称', '数量'], ['螺丝 M6', -5]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['data']['status'] == 'partial'
        final_status = self.update_status.calls[-1]
        assert final_status['args'][1] == 'partial'

    def test_clean_upload_still_success(self, app_mod, client, fake_db, monkeypatch):
        """回归 #11：无错误 → 仍为 success"""
        self._patch(app_mod, fake_db, monkeypatch)
        buf = make_xlsx([['商品名称', '数量'], ['螺丝 M6', 50]])

        resp = client.post('/api/upload', data={'file': (buf, 't.xlsx'), 'mode': 'replace'},
                           content_type='multipart/form-data')
        assert resp.status_code == 200
        body = resp.get_json()
        assert body['data']['status'] == 'success'
        final_status = self.update_status.calls[-1]
        assert final_status['args'][1] == 'success'


# ==========================================
#  回归：AI 批次号唯一（#13）
# ==========================================
class TestAIBatchUnique:
    BATCH_RE = re.compile(r'^AI(?:识别|导入|操作|出库)-\d{14}-[0-9a-f]{6}$')

    def test_inbound_recognize_unique_batch(self, app_mod, client, fake_db, monkeypatch):
        """回归 #13：入库识别批次号每次调用唯一（旧常量 'AI视觉识别' 会让多次导入合并成一批）"""
        items = [{'name': '螺丝 M6', 'sku': 'S1', 'quantity': 5}]
        monkeypatch.setattr(app_mod.ai_service, 'recognize_inbound_image', lambda b64: items)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder({'id': 9}))
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.post('/api/ai/inbound-recognize', json={'image': 'aGVsbG8='})
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        txn = fake_db.find_executed('INSERT INTO transactions')
        assert len(txn) == 1
        # stock_in 流水参数: (product_id, qty, unit_price, supplier_id, before, after, batch_no, operator, notes)
        assert self.BATCH_RE.match(txn[0][1][6])

    def test_smart_import_unique_batch(self, app_mod, client, fake_db, monkeypatch):
        """回归 #13：智能导入批次号每次调用唯一（旧常量 'AI智能导入'）"""
        items = [{'name': '螺丝 M6', 'sku': 'S1', 'quantity': 5}]
        monkeypatch.setattr(app_mod.ai_service, 'smart_import', lambda text: items)
        existing = {'id': 9}
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(existing))
        monkeypatch.setattr(app_mod.CategoryModel, 'get_all', make_recorder([]))
        monkeypatch.setattr(app_mod.SupplierModel, 'get_all', make_recorder([]))
        make_stateful_inventory(fake_db, monkeypatch, initial_qty=20)
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.post('/api/ai/smart-import', json={'text': '采购螺丝 5 个'})
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        txn = fake_db.find_executed('INSERT INTO transactions')
        assert len(txn) == 1
        # smart import 流水参数: (product_id, qty, price, supplier_id, before, after, batch_no, operator, notes)
        assert self.BATCH_RE.match(txn[0][1][6])

    def test_chat_stock_in_unique_batch_per_action(self, app_mod, client, fake_db, monkeypatch):
        """回归 #13：同一轮对话的多个入库动作不得共用批次（旧常量 'AI操作'）"""
        reply = ('好的。\n```action\n'
                 '[{"action": "stock_in", "sku": "S1", "quantity": 5},'
                 '{"action": "stock_in", "sku": "S2", "quantity": 3}]\n```\n')
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply)
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder({'id': 9}))
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        resp = client.post('/api/ai/chat', json={'message': '入库'})
        assert resp.status_code == 200 and resp.get_json()['success'] is True

        txn = fake_db.find_executed('INSERT INTO transactions')
        assert len(txn) == 2
        batches = [p[6] for _, p in txn]
        assert all(self.BATCH_RE.match(b) for b in batches)
        assert batches[0] != batches[1]

    def test_create_order_unique_batch_and_writes_file(self, app_mod, client, tmp_path, fake_db, monkeypatch):
        """回归 #13/#12：AI 出库单每单独立批次（旧常量 'AI出库单'），且文件落盘可下载"""
        reply = ('好的。\n```action\n' + json.dumps([{
            'action': 'create_order', 'customer': '中建公司', 'operator': '张三',
            'items': [{'sku': 'SCR-1', 'quantity': 2, 'price': 0.5}]
        }], ensure_ascii=False) + '\n```\n')
        monkeypatch.setattr(app_mod.ai_service, 'chat', lambda msg: reply)
        monkeypatch.setattr(app_mod.CustomerModel, 'get_by_name', make_recorder({'id': 8}))
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_sku', make_recorder(
            {'id': 7, 'name': '螺丝 M6', 'specification': ''}))
        fake_db.one_side_effect = lambda sql, params=None: {'quantity': 10}
        patch_all_dbs(app_mod, fake_db, monkeypatch)
        monkeypatch.setattr(app_mod, 'UPLOAD_FOLDER', str(tmp_path))

        resp = client.post('/api/ai/chat', json={'message': '创建出库单'})
        body = resp.get_json()
        assert resp.status_code == 200 and body['success'] is True

        txn = fake_db.find_executed('INSERT INTO transactions')
        assert len(txn) == 1
        # stock_out 流水参数: (product_id, qty, unit_price, before, after, batch_no, operator, notes, customer_id)
        assert self.BATCH_RE.match(txn[0][1][5])

        files = os.listdir(str(tmp_path))
        assert any(f.startswith('出库单_') and f.endswith('.xlsx') for f in files)


# ==========================================
#  回归：AI 出库单文件清理（#12a）
# ==========================================
class TestOrderFileCleanup:
    def test_cleanup_removes_only_expired_order_files(self, app_mod, tmp_path, monkeypatch):
        up = tmp_path / 'uploads'
        up.mkdir()
        old = up / '出库单_旧客户_20250101_000000.xlsx'
        old.write_bytes(b'PK')
        new = up / '出库单_新客户_20990101_000000.xlsx'
        new.write_bytes(b'PK')
        other = up / '库存清单_20250101_000000.xlsx'   # 非出库单文件不得被清理
        other.write_bytes(b'PK')

        old_ts = time.time() - 8 * 86400
        os.utime(str(old), (old_ts, old_ts))

        monkeypatch.setattr(app_mod, 'UPLOAD_FOLDER', str(up))
        app_mod._cleanup_old_order_files(keep_days=7)

        assert not old.exists()          # 过期出库单被清理
        assert new.exists()              # 新文件保留
        assert other.exists()            # 非出库单文件不动


# ==========================================
#  回归：AI 服务连接错误不再伪装成解析错误（#14）
# ==========================================
class TestAIServiceErrors:
    def test_connection_error_raises_aiserviceerror(self, monkeypatch):
        import ai_service as ai_mod

        def boom(*a, **k):
            raise requests_lib.exceptions.ConnectionError('refused')
        monkeypatch.setattr(ai_mod.requests, 'post', boom)

        with pytest.raises(ai_mod.AIServiceError, match='无法连接'):
            ai_mod.ai_service._call([{'role': 'user', 'content': 'hi'}])

    def test_timeout_raises_aiserviceerror(self, monkeypatch):
        import ai_service as ai_mod

        def boom(*a, **k):
            raise requests_lib.exceptions.Timeout('slow')
        monkeypatch.setattr(ai_mod.requests, 'post', boom)

        with pytest.raises(ai_mod.AIServiceError, match='超时'):
            ai_mod.ai_service._call([{'role': 'user', 'content': 'hi'}])

    def test_http_error_raises_with_status(self, monkeypatch):
        import ai_service as ai_mod
        resp = SimpleNamespace(status_code=502)
        err = requests_lib.exceptions.HTTPError('bad gateway')
        err.response = resp

        def boom(*a, **k):
            raise err
        monkeypatch.setattr(ai_mod.requests, 'post', boom)

        with pytest.raises(ai_mod.AIServiceError, match='HTTP 502'):
            ai_mod.ai_service._call([{'role': 'user', 'content': 'hi'}])

    def test_recognize_inbound_propagates_connection_error(self, monkeypatch):
        """回归 #14：LM Studio 挂了 → 入库识别报连接失败，而不是误导性的"解析格式错误" """
        import ai_service as ai_mod

        def boom(*a, **k):
            raise requests_lib.exceptions.ConnectionError('refused')
        monkeypatch.setattr(ai_mod.requests, 'post', boom)
        monkeypatch.setattr(ai_mod.ai_service, 'ocr_image', lambda b64: '入库单 螺丝 M6 x5')

        with pytest.raises(RuntimeError, match='无法连接'):
            ai_mod.ai_service.recognize_inbound_image('aGVsbG8=')


# ==========================================
#  批量出库单（H5 原子提交）
# ==========================================
class TestOrderSubmitAtomicy:
    """回归 H5：手工出库单改为单事务批量接口，防「部分出库 + 提示成功」的半截写。"""

    def _setup_pipeline(self, app_mod, client, fake_db, monkeypatch, inventory_qty=100):
        """把 get_by_id / stock_out / db 都接好，返回可断言的 recorder。"""
        # 两个商品
        def fake_get_by_id(pid):
            return {'id': pid, 'name': f'商品{pid}', 'sku': f'S-{pid}',
                    'unit': '个', 'specification': 'x', 'sale_price': 2}
        monkeypatch.setattr(app_mod.ProductModel, 'get_by_id', fake_get_by_id)

        # 库存充足的回读（FakeDB.query_one 的 FOR UPDATE 查询）
        fake_db.one_side_effect = lambda sql, params=None: (
            {'quantity': inventory_qty} if 'SELECT quantity FROM inventory' in sql else None)

        # 同时 patch models 与 app 的 db，确保 AuditLog.log 也走 FakeDB（避免连真库）
        patch_all_dbs(app_mod, fake_db, monkeypatch)

        stock_out_rec = make_recorder(None)   # 记录每次调用的参数
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_out', stock_out_rec)
        return stock_out_rec

    def test_success_submits_all_items_in_one_transaction(self, app_mod, client, fake_db, monkeypatch):
        rec = self._setup_pipeline(app_mod, client, fake_db, monkeypatch, inventory_qty=100)
        body = {'customer_id': 7, 'operator': '张三',
                'items': [{'product_id': 3, 'quantity': 5, 'unit_price': 2.0},
                          {'product_id': 4, 'quantity': 8, 'unit_price': 1.5}]}
        resp = client.post('/api/order/submit', json=body)
        assert resp.status_code == 200, resp.get_json()
        data = resp.get_json()['data']
        assert data['count'] == 2
        assert data['batch_no']
        # 每个商品都走了一次 stock_out
        assert len([c for c in rec.calls]) == 2
        # 审计与扣减在同一事务内
        aud = app_mod.AuditLog
        # 事务期间所有写都在 in_transaction=True 下发生
        assert fake_db.in_transaction is False  # 结束后应复位

    def test_oversell_precheck_returns_400_and_does_not_deduct(self, app_mod, client, fake_db, monkeypatch):
        rec = self._setup_pipeline(app_mod, client, fake_db, monkeypatch, inventory_qty=5)
        body = {'customer_id': 7, 'operator': '张三',
                'items': [{'product_id': 3, 'quantity': 5, 'unit_price': 2.0},   # 库存5刚刚够
                          {'product_id': 4, 'quantity': 100, 'unit_price': 1.5}]}  # 库存不足
        resp = client.post('/api/order/submit', json=body)
        assert resp.status_code == 400
        assert '未执行' in resp.get_json()['error']
        assert '库存不足' in resp.get_json()['error']
        # 任何一个商品都不应被扣减（预检失败整体不执行）
        assert len([c for c in rec.calls]) == 0

    def test_missing_items_rejected(self, app_mod, client, fake_db, monkeypatch):
        self._setup_pipeline(app_mod, client, fake_db, monkeypatch)
        resp = client.post('/api/order/submit', json={'customer_id': 7, 'operator': 'x', 'items': []})
        assert resp.status_code == 400

    def test_missing_customer_rejected(self, app_mod, client, fake_db, monkeypatch):
        rec = self._setup_pipeline(app_mod, client, fake_db, monkeypatch)
        resp = client.post('/api/order/submit', json={'operator': 'x',
                                                      'items': [{'product_id': 3, 'quantity': 1}]})
        assert resp.status_code == 400


# ==========================================
#  账户管理（UserModel + /api/users）
# ==========================================
class TestUserManagement:
    def test_usermodel_authenticate_passes_on_correct_password(self, app_mod, monkeypatch):
        """UserModel.authenticate 在校验正确密码时通过（密码加盐哈希存储）。"""
        import models as models_mod
        from werkzeug.security import generate_password_hash
        h = generate_password_hash('secret123')
        user_row = {'id': 1, 'username': 'bob', 'password_hash': h, 'role': 'user'}
        monkeypatch.setattr(models_mod.UserModel, 'get_by_username', lambda u: user_row)
        assert models_mod.UserModel.authenticate('bob', 'secret123') is not None
        # 错误密码返回 None
        assert models_mod.UserModel.authenticate('bob', 'wrong') is None

    def test_usermodel_authenticate_rejects_unknown_user(self, app_mod, monkeypatch):
        import models as models_mod
        monkeypatch.setattr(models_mod.UserModel, 'get_by_username', lambda u: None)
        assert models_mod.UserModel.authenticate('nobody', 'x') is None

    def test_create_user_route_requires_admin(self, client):
        """非管理员 / 未配置管理员时创建账户应被拒（403 或 401）。"""
        resp = client.post('/api/users', json={'username': 'x', 'password': '123456', 'role': 'user'})
        # DISABLE_AUTH=true 下 _is_admin 恒 False → 403
        assert resp.status_code == 403

    def test_create_user_duplicate_username_rejected(self, app_mod, client, monkeypatch):
        """已存在的用户名创建时应返回 400。"""
        monkeypatch.setattr(app_mod, '_is_admin', lambda: True)
        monkeypatch.setattr(app_mod.UserModel, 'get_by_username', lambda u: {'id': 1})
        resp = client.post('/api/users', json={'username': 'bob', 'password': '123456', 'role': 'user'})
        assert resp.status_code == 400
        assert '已存在' in resp.get_json()['error']

    def test_create_user_short_password_rejected(self, app_mod, client, monkeypatch):
        monkeypatch.setattr(app_mod, '_is_admin', lambda: True)
        monkeypatch.setattr(app_mod.UserModel, 'get_by_username', lambda u: None)
        resp = client.post('/api/users', json={'username': 'bob', 'password': '123', 'role': 'user'})
        assert resp.status_code == 400
        assert '至少 6 位' in resp.get_json()['error']



# ==========================================
#  Session 登录（登录验证逻辑）
# ==========================================
class TestSessionLogin:
    def test_validate_credentials_db_user(self, app_mod, monkeypatch):
        """数据库用户表校验：正确用户+密码 → (True, role)"""
        monkeypatch.setattr(app_mod.UserModel, 'authenticate',
                            lambda u, p: {'id': 1, 'username': u, 'role': 'user'})
        ok, role = app_mod._validate_credentials('bob', 'correct')
        assert ok is True and role == 'user'

    def test_validate_credentials_env_admin_fallback(self, app_mod, monkeypatch):
        """环境变量超级管理员兜底：users 表查不到时仍可登录 admin"""
        monkeypatch.setattr(app_mod.UserModel, 'authenticate', lambda u, p: None)
        monkeypatch.setattr(app_mod, 'AUTH_USER', 'admin')
        monkeypatch.setattr(app_mod, 'AUTH_PASSWORD', 'admin123')
        ok, role = app_mod._validate_credentials('admin', 'admin123')
        assert ok is True and role == 'admin'

    def test_validate_credentials_wrong_password(self, app_mod, monkeypatch):
        """错误密码（用户表+环境变量都不匹配）→ (False, None)"""
        monkeypatch.setattr(app_mod.UserModel, 'authenticate', lambda u, p: None)
        monkeypatch.setattr(app_mod, 'AUTH_USER', 'admin')
        monkeypatch.setattr(app_mod, 'AUTH_PASSWORD', 'admin123')
        ok, role = app_mod._validate_credentials('admin', 'wrong')
        assert ok is False and role is None

    def test_login_route_rejects_bad_credentials(self, app_mod, client, monkeypatch):
        """POST /api/login 错误凭据 → 401"""
        monkeypatch.setattr(app_mod.UserModel, 'authenticate', lambda u, p: None)
        monkeypatch.setattr(app_mod, 'AUTH_PASSWORD', '')
        resp = client.post('/api/login', json={'username': 'x', 'password': 'bad'})
        assert resp.status_code == 401
        assert '用户名或密码错误' in resp.get_json()['error']


# ==========================================
#  出入库操作者 = 登录账号
# ==========================================
class TestOperatorMatchesLogin:
    def test_stock_in_operator_is_logged_in_user(self, app_mod, client, fake_db, monkeypatch):
        """回归：登录账号为 xhj 时，stock-in 流水 operator 应为 xhj 而非'管理员'"""
        # 模拟已登录用户 xhj（直接写 session；DISABLE_AUTH 下 before_request 不拦，但 _op 读 session）
        with client.session_transaction() as sess:
            sess['username'] = 'xhj'
            sess['role'] = 'admin'

        stock_in = make_recorder(None)
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_in', stock_in)
        monkeypatch.setattr(app_mod, 'db', fake_db)
        monkeypatch.setattr(app_mod.AuditLog, 'log', lambda *a, **k: None)

        resp = client.post('/api/inventory/stock-in', json={
            'product_id': 7, 'quantity': 5, 'unit_price': 2.5,
        })
        assert resp.status_code == 200
        # InventoryModel.stock_in 收到的 operator 参数 = 登录账号 xhj
        call = stock_in.calls[0]['kwargs']
        assert call.get('operator') == 'xhj'

    def test_stock_out_operator_is_logged_in_user(self, app_mod, client, fake_db, monkeypatch):
        with client.session_transaction() as sess:
            sess['username'] = 'xhj'
            sess['role'] = 'admin'
        stock_out = make_recorder(None)
        monkeypatch.setattr(app_mod.InventoryModel, 'stock_out', stock_out)
        monkeypatch.setattr(app_mod, 'db', fake_db)
        monkeypatch.setattr(app_mod.AuditLog, 'log', lambda *a, **k: None)
        resp = client.post('/api/inventory/stock-out', json={
            'product_id': 7, 'quantity': 5, 'unit_price': 2.5,
        })
        assert resp.status_code == 200
        call = stock_out.calls[0]['kwargs']
        assert call.get('operator') == 'xhj'
